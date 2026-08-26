# 4C大赛参赛作品：膳康管家 - 个性化膳食图像识别与营养管理系统
# 开发：AI专业参赛团队
# 合规说明：全程使用大赛指定国产AI工具开发，符合赛事要求
import gradio as gr
import torch
import datetime
import json
import os
import io
import logging
import tempfile
import hashlib
import secrets
import shutil
import uuid
from ollama import Client  # 不要直接 import ollama，而是引入 Client 类
from ultralytics import YOLO
from PIL import Image, ImageDraw, ImageFont
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
import ollama  # 用于调用本地 Ollama 模型
import threading  # 用于线程安全的文件读写锁
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

# -------------------------- 0. 日志系统初始化 --------------------------
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler('app.log', encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# 启动时检查关键依赖
logger.info("正在初始化膳康管家系统...")
try:
    import openpyxl
    logger.info("openpyxl 已就绪，Excel 导出功能可用")
except ImportError:
    logger.warning("openpyxl 未安装，Excel 导出功能将受限。请运行: pip install openpyxl")

# 1. 暴力清除当前 Python 进程内的所有代理环境变量（比 NO_PROXY 更彻底）
os.environ['HTTP_PROXY'] = ''
os.environ['HTTPS_PROXY'] = ''
os.environ['ALL_PROXY'] = ''
os.environ['NO_PROXY'] = 'localhost,127.0.0.1,::1'

# 2. 显式创建一个指向纯 IPv4 本地地址的客户端，彻底绕过 localhost 解析
ollama_client = Client(host='http://127.0.0.1:11434')

# -------------------------- 1. 全局配置与初始化 --------------------------
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
logger.info(f"当前运行设备：{device}")

# 配置 Matplotlib 中文字体，防止乱码（豆腐块）
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'Arial Unicode MS']
plt.rcParams['axes.unicode_minus'] = False

# 头像图片目录（绝对路径，保证跨环境兼容）
PICTURE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "picture")

MODEL_PATH = r"model\best.pt"
model = YOLO(MODEL_PATH)
model.eval()

FOOD_CLASSES = [
    # 0~5 必须是原来的水果，顺序不能乱
    "苹果", "香蕉", "葡萄", "橙子", "菠萝", "西瓜",

    # 6~63 必须严格按照 YAML 文件里的英文顺序对应的中文
    "咸蛋苦瓜", "黑椒牛柳", "卤蛋", "红烧肉", "凉拌黄瓜", "茄子豆角",
    "炸鸡腿", "炒茄子", "炒青椒", "炒面", "炒猪血", "炸排骨",
    "炒饭", "炒米粉", "炒地瓜叶", "炒青豆", "炒小白菜", "客家小炒",
    "宫保鸡丁", "煎猪肉", "土豆鸡蛋沙拉", "番茄炒蛋", "炒鸡蛋", "炒肉丝",
    "三色蒸蛋", "甜辣炒韭菜花", "海带炒豆干", "小鱼干炒豆干", "西葫芦", "芦笋",
    "红烧鱼", "胡萝卜炒蛋", "鸡腿", "炸鸡块", "大白菜", "腊肠",
    "玉米", "咖喱", "炒笋片", "炒四季豆", "炸鸡", "煎饺",
    "煎蛋", "海带", "丝瓜", "绿豆芽", "秋葵", "南瓜",
    "米饭", "三文鱼", "芝麻豆腐", "虾", "小鱼", "蒸蛋",
    "炒黄瓜球", "红薯", "三角薯饼", "空心菜"
]

NUTRITION_DB = {
    # 0~5 水果 (原始提供数据)
    "苹果": [52, 0.2, 13.5, 0.2, 1.7, 1.3, 36],
    "香蕉": [91, 1.4, 22.0, 0.2, 1.2, 0.8, 52],
    "葡萄": [43, 0.5, 10.3, 0.2, 0.4, 1.3, 40],
    "橙子": [47, 0.8, 11.1, 0.2, 1.4, 1.2, 40],
    "菠萝": [44, 0.5, 10.8, 0.1, 1.3, 0.8, 66],
    "西瓜": [31, 0.6, 7.6, 0.1, 0.3, 1.8, 72],

    # 6~63 菜肴与食材
    "咸蛋苦瓜": [110, 4.0, 5.0, 8.0, 2.0, 400, 45],
    "黑椒牛柳": [160, 18.0, 6.0, 7.0, 1.0, 500, 45],
    "卤蛋": [155, 13.0, 3.0, 10.0, 0.0, 600, 30],
    "红烧肉": [470, 12.0, 8.0, 43.0, 0.0, 550, 45],
    "凉拌黄瓜": [45, 1.0, 3.0, 3.0, 1.0, 450, 15],
    "茄子豆角": [90, 2.0, 6.0, 6.0, 3.0, 350, 35],

    "炸鸡腿": [260, 16.0, 10.0, 17.0, 0.0, 350, 55],
    "炒茄子": [110, 1.0, 5.0, 9.0, 2.0, 300, 35],
    "炒青椒": [70, 1.0, 4.0, 5.0, 2.0, 250, 15],
    "炒面": [180, 4.0, 25.0, 6.0, 2.0, 400, 65],
    "炒猪血": [80, 8.0, 2.0, 4.0, 0.0, 300, 20],
    "炸排骨": [320, 15.0, 8.0, 25.0, 0.0, 450, 50],

    "炒饭": [180, 4.0, 26.0, 6.0, 1.0, 350, 70],
    "炒米粉": [190, 3.0, 28.0, 7.0, 1.0, 350, 65],
    "炒地瓜叶": [65, 2.0, 3.0, 5.0, 3.0, 250, 15],
    "炒青豆": [110, 6.0, 8.0, 6.0, 4.0, 250, 30],
    "炒小白菜": [55, 1.0, 2.0, 4.0, 1.0, 200, 15],
    "客家小炒": [210, 14.0, 4.0, 15.0, 1.0, 550, 40],

    "宫保鸡丁": [190, 14.0, 8.0, 11.0, 1.0, 450, 45],
    "煎猪肉": [280, 16.0, 2.0, 22.0, 0.0, 300, 0],
    "土豆鸡蛋沙拉": [150, 4.0, 12.0, 10.0, 1.5, 250, 65],
    "番茄炒蛋": [110, 5.0, 5.0, 8.0, 1.0, 350, 45],
    "炒鸡蛋": [160, 12.0, 2.0, 11.0, 0.0, 250, 30],
    "炒肉丝": [220, 15.0, 3.0, 16.0, 0.0, 400, 35],

    "三色蒸蛋": [110, 10.0, 2.0, 7.0, 0.0, 450, 30],
    "甜辣炒韭菜花": [80, 3.0, 6.0, 5.0, 3.0, 300, 35],
    "海带炒豆干": [120, 8.0, 5.0, 7.0, 3.0, 400, 30],
    "小鱼干炒豆干": [180, 15.0, 4.0, 11.0, 2.0, 600, 30],
    "西葫芦": [19, 1.0, 4.0, 0.2, 1.0, 2.0, 15],
    "芦笋": [20, 2.0, 4.0, 0.1, 2.0, 2.0, 15],

    "红烧鱼": [140, 15.0, 4.0, 7.0, 0.0, 450, 35],
    "胡萝卜炒蛋": [120, 6.0, 6.0, 8.0, 2.0, 300, 45],
    "鸡腿": [160, 18.0, 0.0, 9.0, 0.0, 80.0, 0],
    "炸鸡块": [290, 14.0, 15.0, 20.0, 1.0, 500, 65],
    "大白菜": [16, 1.0, 3.0, 0.2, 1.0, 10.0, 15],
    "腊肠": [500, 14.0, 10.0, 45.0, 0.0, 1200, 45],

    "玉米": [110, 3.0, 23.0, 1.0, 3.0, 5.0, 55],
    "咖喱": [130, 6.0, 8.0, 8.0, 1.0, 350, 55],
    "炒笋片": [60, 2.0, 5.0, 4.0, 2.0, 250, 20],
    "炒四季豆": [85, 2.0, 6.0, 6.0, 3.0, 250, 30],
    "炸鸡": [270, 15.0, 12.0, 18.0, 0.0, 400, 65],
    "煎饺": [240, 8.0, 25.0, 12.0, 1.5, 400, 60],

    "煎蛋": [190, 13.0, 1.0, 14.0, 0.0, 250, 30],
    "海带": [15, 1.0, 3.0, 0.1, 2.0, 50.0, 15],
    "丝瓜": [20, 1.0, 4.0, 0.2, 1.0, 2.0, 15],
    "绿豆芽": [16, 2.0, 3.0, 0.1, 1.0, 2.0, 25],
    "秋葵": [33, 2.0, 7.0, 0.1, 3.0, 7.0, 20],
    "南瓜": [26, 1.0, 6.0, 0.1, 1.0, 1.0, 75],

    "米饭": [116, 2.6, 26.0, 0.3, 0.3, 1.0, 83],
    "三文鱼": [139, 20.0, 0.0, 6.0, 0.0, 50.0, 0],
    "芝麻豆腐": [120, 6.0, 5.0, 8.0, 1.0, 150, 35],
    "虾": [95, 20.0, 0.0, 1.0, 0.0, 150, 0],
    "小鱼": [100, 18.0, 0.0, 2.0, 0.0, 100, 0],
    "蒸蛋": [60, 5.0, 1.0, 4.0, 0.0, 150, 30],

    "炒黄瓜球": [50, 1.0, 3.0, 4.0, 1.0, 200, 15],
    "红薯": [90, 1.5, 20.0, 0.2, 3.0, 40.0, 55],
    "三角薯饼": [260, 3.0, 28.0, 15.0, 2.0, 350, 75],
    "空心菜": [20, 2.0, 3.0, 0.2, 2.0, 10.0, 15]
}

# 每种食物默认单份重量（克），用于无称重时的营养估算
DEFAULT_PORTION_WEIGHT = {
    # 水果类（0~5）
    "苹果": 200, "香蕉": 120, "葡萄": 150, "橙子": 180, "菠萝": 250, "西瓜": 500,
    # 主食类
    "米饭": 150, "炒饭": 250, "炒面": 250, "炒米粉": 200, "红薯": 200,
    # 荤菜类
    "咸蛋苦瓜": 200, "黑椒牛柳": 200, "卤蛋": 50, "红烧肉": 200, "炸鸡腿": 120,
    "炒猪血": 150, "炸排骨": 180, "客家小炒": 200, "宫保鸡丁": 200, "煎猪肉": 150,
    "炒肉丝": 180, "红烧鱼": 200, "鸡腿": 120, "炸鸡块": 150, "炸鸡": 150,
    "三文鱼": 120, "虾": 100, "小鱼": 80, "腊肠": 80, "煎饺": 150,
    # 素菜/蛋类
    "凉拌黄瓜": 150, "茄子豆角": 200, "炒茄子": 200, "炒青椒": 150,
    "炒地瓜叶": 150, "炒青豆": 120, "炒小白菜": 150, "炒鸡蛋": 120,
    "番茄炒蛋": 200, "胡萝卜炒蛋": 180, "煎蛋": 100, "蒸蛋": 100,
    "三色蒸蛋": 150, "土豆鸡蛋沙拉": 200, "炒笋片": 150, "炒四季豆": 150,
    "炒黄瓜球": 150, "空心菜": 150, "大白菜": 150, "西葫芦": 200,
    "芦笋": 120, "丝瓜": 200, "绿豆芽": 100, "秋葵": 100, "南瓜": 200,
    "玉米": 200, "海带": 100, "芝麻豆腐": 150, "咖喱": 250,
    "甜辣炒韭菜花": 150, "海带炒豆干": 180, "小鱼干炒豆干": 180,
    "三角薯饼": 100,
}

DATA_DIR = "user_data"
os.makedirs(DATA_DIR, exist_ok=True)

# -------------------------- 1.1 全局营养数据加载 --------------------------
def _try_load_json_data():
    """启动时尝试从JSON加载营养数据，失败则静默使用硬编码兜底"""
    nutrition_json = os.path.join(DATA_DIR, "nutrition_db.json")
    if os.path.exists(nutrition_json):
        try:
            with open(nutrition_json, 'r', encoding='utf-8') as f:
                loaded = json.load(f)
            NUTRITION_DB.update(loaded)
            logger.info(f"已从JSON加载营养数据库（{len(loaded)} 条）")
        except Exception as e:
            logger.warning(f"营养数据库JSON加载失败，使用硬编码兜底: {e}")

    portion_json = os.path.join(DATA_DIR, "portion_weight.json")
    if os.path.exists(portion_json):
        try:
            with open(portion_json, 'r', encoding='utf-8') as f:
                loaded = json.load(f)
            DEFAULT_PORTION_WEIGHT.update(loaded)
            logger.info(f"已从JSON加载默认份量表（{len(loaded)} 条）")
        except Exception as e:
            logger.warning(f"默认份量表JSON加载失败，使用硬编码兜底: {e}")

_try_load_json_data()

# -------------------------- 1.2 旧数据备份 --------------------------
def _backup_old_data():
    """将旧版单用户数据文件移至备份目录，避免干扰多用户系统"""
    old_files = [
        os.path.join(DATA_DIR, "records.json"),
        os.path.join(DATA_DIR, "profile.json"),
    ]
    backup_dir = os.path.join(DATA_DIR, "_old_backup")
    for f in old_files:
        if os.path.exists(f):
            os.makedirs(backup_dir, exist_ok=True)
            dest = os.path.join(backup_dir, os.path.basename(f))
            try:
                shutil.move(f, dest)
                logger.info(f"旧数据文件已备份：{f} → {dest}")
            except Exception as e:
                logger.warning(f"旧数据备份失败 {f}: {e}")

_backup_old_data()

# -------------------------- 2. 用户管理系统 --------------------------
USERS_FILE = os.path.join(DATA_DIR, "users.json")
_users_lock = threading.Lock()

def _load_users():
    """加载所有注册用户信息"""
    with _users_lock:
        if os.path.exists(USERS_FILE):
            try:
                with open(USERS_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except (json.JSONDecodeError, IOError) as e:
                logger.error(f"用户文件读取失败: {e}")
                return {}
        return {}

def _save_users(users_dict):
    """保存用户信息到文件"""
    with _users_lock:
        # 写入前备份
        if os.path.exists(USERS_FILE):
            try:
                shutil.copy2(USERS_FILE, USERS_FILE + ".bak")
            except Exception as e:
                logger.warning(f"用户文件备份失败: {e}")
        with open(USERS_FILE, 'w', encoding='utf-8') as f:
            json.dump(users_dict, f, ensure_ascii=False, indent=2)

def hash_password(password, salt=None):
    """
    使用 SHA-256 + 随机盐对密码进行哈希
    返回 (password_hash_hex, salt_hex)
    """
    if salt is None:
        salt = secrets.token_hex(16)  # 128位随机盐
    salted = salt + password
    hash_bytes = hashlib.sha256(salted.encode('utf-8')).digest()
    # 多轮哈希增强安全性
    for _ in range(1000):
        hash_bytes = hashlib.sha256(hash_bytes + salted.encode('utf-8')).digest()
    return hash_bytes.hex(), salt

def verify_password(password, salt, stored_hash):
    """验证密码是否匹配"""
    computed_hash, _ = hash_password(password, salt)
    return secrets.compare_digest(computed_hash, stored_hash)

def register_user(username, password):
    """
    注册新用户
    返回 (success: bool, message: str)
    """
    # 校验用户名
    if not username or not username.strip():
        return False, "❌ 用户名不能为空"
    username = username.strip()
    if len(username) < 2:
        return False, "❌ 用户名至少需要2个字符"
    if not username.isalnum():
        return False, "❌ 用户名只能包含字母和数字"

    # 校验密码
    if not password or len(password) < 6:
        return False, "❌ 密码长度至少为6位"

    # 检查用户名是否已存在
    users = _load_users()
    if username in users:
        return False, "❌ 该用户名已被注册，请换一个"

    # 创建用户
    password_hash, salt = hash_password(password)
    users[username] = {
        "password_hash": password_hash,
        "salt": salt,
        "created_at": datetime.datetime.now().isoformat()
    }
    _save_users(users)

    # 为用户创建数据目录
    user_dir = os.path.join(DATA_DIR, username)
    os.makedirs(user_dir, exist_ok=True)

    logger.info(f"新用户注册成功：{username}")
    return True, f"✅ 注册成功，欢迎 {username}！"

def verify_login(username, password):
    """
    验证登录
    返回 (success: bool, message: str)
    """
    if not username or not password:
        return False, "❌ 用户名和密码不能为空"

    users = _load_users()
    if username not in users:
        return False, "❌ 用户名不存在，请先注册"

    user_info = users[username]
    if not verify_password(password, user_info["salt"], user_info["password_hash"]):
        return False, "❌ 密码错误"

    # 确保用户数据目录存在（兼容旧数据迁移场景）
    user_dir = os.path.join(DATA_DIR, username)
    os.makedirs(user_dir, exist_ok=True)

    logger.info(f"用户登录成功：{username}")
    return True, f"✅ 登录成功，欢迎回来 {username}！"

# -------------------------- 3. DietManager 类（多用户重构） --------------------------

class DietManager:
    """膳食数据管理器 —— 每个用户独立一个实例，数据存储在其专属子目录下"""
    _instances_lock = threading.Lock()
    _instances = {}  # 类级别缓存：username -> DietManager

    def __init__(self, username):
        self.username = username
        self.user_dir = os.path.join(DATA_DIR, username)
        os.makedirs(self.user_dir, exist_ok=True)
        self._records_file = os.path.join(self.user_dir, "records.json")
        self._profile_file = os.path.join(self.user_dir, "profile.json")
        self._lock = threading.Lock()
        self.records = self._load_records()
        self.profile = self._load_profile()

    @classmethod
    def get_instance(cls, username):
        """获取或创建用户专属的 DietManager 实例（线程安全）"""
        if not username:
            return None
        with cls._instances_lock:
            if username not in cls._instances:
                cls._instances[username] = cls(username)
            return cls._instances[username]

    def _load_records(self):
        with self._lock:
            if os.path.exists(self._records_file):
                try:
                    with open(self._records_file, 'r', encoding='utf-8') as f:
                        return json.load(f)
                except (json.JSONDecodeError, IOError) as e:
                    logger.error(f"加载膳食记录失败 ({self.username}): {e}")
                    return []
            return []

    def _backup_file(self, filepath):
        """写入前自动备份旧文件为 .bak"""
        if os.path.exists(filepath):
            backup_path = filepath + ".bak"
            try:
                shutil.copy2(filepath, backup_path)
            except Exception as e:
                logger.warning(f"文件备份失败 {filepath}: {e}")

    def _save_records(self):
        with self._lock:
            self._backup_file(self._records_file)
            with open(self._records_file, 'w', encoding='utf-8') as f:
                json.dump(self.records, f, ensure_ascii=False, indent=2)

    def _load_profile(self):
        with self._lock:
            if os.path.exists(self._profile_file):
                try:
                    with open(self._profile_file, 'r', encoding='utf-8') as f:
                        return json.load(f)
                except (json.JSONDecodeError, IOError) as e:
                    logger.error(f"加载个人档案失败 ({self.username}): {e}")
            # 默认档案
            return {
                "name": self.username,
                "height": 170,
                "weight": 65,
                "age": 20,
                "gender": "男",
                "activity": "轻度活动",
                "goal": "维持体重",
                "disease": "健康"
            }

    def _save_profile(self):
        with self._lock:
            self._backup_file(self._profile_file)
            with open(self._profile_file, 'w', encoding='utf-8') as f:
                json.dump(self.profile, f, ensure_ascii=False, indent=2)

    def add_record(self, food_summary, total_weight, total_nutrition, date, meal_type="午餐"):
        weight_ratio = total_weight / 100
        intake = {
            "食品组合": food_summary,
            "总食用重量(g)": total_weight,
            "日期": date,
            "餐次": meal_type,
            "热量(kcal)": round(total_nutrition[0] * weight_ratio, 1),
            "蛋白质(g)": round(total_nutrition[1] * weight_ratio, 1),
            "碳水(g)": round(total_nutrition[2] * weight_ratio, 1),
            "脂肪(g)": round(total_nutrition[3] * weight_ratio, 1),
            "膳食纤维(g)": round(total_nutrition[4] * weight_ratio, 1),
            "钠(mg)": round(total_nutrition[5] * weight_ratio, 1),
            "平均GI值": total_nutrition[6],
            "时间戳": datetime.datetime.now().isoformat()
        }
        self.records.append(intake)
        self._save_records()
        return f"✅ 记录添加成功！共识别 {food_summary}，总重 {total_weight}g ({meal_type})"

    def delete_record(self, index):
        """按索引删除一条膳食记录，索引从0开始"""
        if 0 <= index < len(self.records):
            deleted = self.records.pop(index)
            self._save_records()
            return f"✅ 已删除记录：{deleted.get('食品组合', '未知')} ({deleted.get('日期', '')} {deleted.get('餐次', '')})"
        return "❌ 删除失败：索引超出范围，请检查记录序号"

    def get_daily_stats(self, date):
        std_date = _standardize_date(date)
        daily_records = [r for r in self.records if r["日期"] == std_date]
        if not daily_records:
            return "当日暂无膳食记录", None, None, None

        total_intake = {
            "热量(kcal)": 0, "蛋白质(g)": 0, "碳水(g)": 0,
            "脂肪(g)": 0, "膳食纤维(g)": 0, "钠(mg)": 0
        }
        for record in daily_records:
            for key in total_intake.keys():
                total_intake[key] += record[key]

        recommend = self.calculate_recommend_intake()
        report = self._generate_report(std_date, total_intake, recommend, daily_records)
        progress_img = self._create_progress_rings(total_intake, recommend)
        return report, pd.DataFrame(daily_records), total_intake, progress_img

    def get_weekly_stats(self):
        today = datetime.date.today()
        week_records = []
        for i in range(7):
            date_str = (today - datetime.timedelta(days=i)).isoformat()
            daily = [r for r in self.records if r["日期"] == date_str]
            if daily:
                week_records.extend(daily)

        if not week_records:
            return "本周暂无记录", None

        df = pd.DataFrame(week_records)
        chart = self._create_weekly_chart(df)
        return f"本周共记录 {len(week_records)} 餐", chart

    def calculate_recommend_intake(self):
        p = self.profile
        height, weight, age, gender = p["height"], p["weight"], p["age"], p["gender"]
        activity, goal = p["activity"], p["goal"]

        if gender == "男":
            bmr = 10 * weight + 6.25 * height - 5 * age + 5
        else:
            bmr = 10 * weight + 6.25 * height - 5 * age - 161

        activity_coeff = {
            "久坐不动": 1.2, "轻度活动": 1.375, "中度活动": 1.55,
            "重度活动": 1.725, "极重度活动": 1.9
        }[activity]

        tdee = bmr * activity_coeff
        goal_adjust = {
            "减脂": 0.85, "维持体重": 1.0, "增肌": 1.15,
            "糖尿病管理": 0.9, "高血压管理": 0.95
        }[goal]

        recommend_calorie = tdee * goal_adjust
        recommend_protein = weight * (1.2 if goal != "增肌" else 1.8)
        recommend_carbs = (recommend_calorie * 0.5) / 4
        recommend_fat = (recommend_calorie * 0.25) / 9
        recommend_sodium = 2000 if goal in ["高血压管理"] else 2300

        return {
            "热量": round(recommend_calorie, 1),
            "蛋白质": round(recommend_protein, 1),
            "碳水": round(recommend_carbs, 1),
            "脂肪": round(recommend_fat, 1),
            "钠": round(recommend_sodium, 1)
        }

    def _generate_report(self, date, intake, recommend, daily_records):
        report = f"📅 {date} 膳食营养统计报告\n"
        report += "=" * 50 + "\n"

        # 按餐次维度拆分统计
        meal_groups = {"早餐": [], "午餐": [], "晚餐": [], "加餐": []}
        for r in daily_records:
            meal_groups.get(r.get("餐次", "午餐"), []).append(r)

        report += "📋 各餐次摄入概况：\n"
        for meal, records in meal_groups.items():
            if records:
                meal_cal = sum(r["热量(kcal)"] for r in records)
                meal_protein = sum(r["蛋白质(g)"] for r in records)
                meal_carbs = sum(r["碳水(g)"] for r in records)
                meal_fat = sum(r["脂肪(g)"] for r in records)
                report += f"  {meal}：{meal_cal:.0f} kcal | 蛋白质 {meal_protein:.0f}g | 碳水 {meal_carbs:.0f}g | 脂肪 {meal_fat:.0f}g\n"
            else:
                report += f"  {meal}：暂无记录\n"
        report += "=" * 50 + "\n"

        report += f"🍚 总摄入热量：{intake['热量(kcal)']} kcal / 推荐 {recommend['热量']} kcal "
        report += self._get_progress_bar(intake['热量(kcal)'], recommend['热量']) + "\n"
        report += f"🥚 蛋白质：{intake['蛋白质(g)']} g / 推荐 {recommend['蛋白质']} g "
        report += self._get_progress_bar(intake['蛋白质(g)'], recommend['蛋白质']) + "\n"
        report += f"🍞 碳水化合物：{intake['碳水(g)']} g / 推荐 {recommend['碳水']} g "
        report += self._get_progress_bar(intake['碳水(g)'], recommend['碳水']) + "\n"
        report += f"🥩 脂肪：{intake['脂肪(g)']} g / 推荐 {recommend['脂肪']} g "
        report += self._get_progress_bar(intake['脂肪(g)'], recommend['脂肪']) + "\n"
        report += f"🥬 膳食纤维：{intake['膳食纤维(g)']} g / 推荐 25g "
        report += self._get_progress_bar(intake['膳食纤维(g)'], 25) + "\n"
        report += f"🧂 钠：{intake['钠(mg)']} mg / 推荐 {recommend['钠']} mg "
        report += self._get_progress_bar(intake['钠(mg)'], recommend['钠']) + "\n"
        report += "=" * 50 + "\n"

        if daily_records:
            avg_gi = sum(r['平均GI值'] for r in daily_records) / len(daily_records)
            gi_level = "低" if avg_gi < 55 else "中" if avg_gi < 70 else "高"
            report += f"📊 平均GI值：{avg_gi:.1f} ({gi_level}GI)\n\n"

        report += self.generate_advice(intake, recommend)
        report += "\n" + self._calculate_exercise(intake['热量(kcal)'])
        return report

    def _get_progress_bar(self, current, target, width=20):
        if target == 0:
            return ""
        ratio = min(current / target, 1.0)
        filled = int(width * ratio)
        bar = "█" * filled + "░" * (width - filled)
        percentage = (ratio * 100)
        color = "🟢" if ratio <= 1.0 else "🔴"
        return f"[{bar}] {percentage:.0f}% {color}"

    def _calculate_exercise(self, calories):
        exercises = {
            "慢跑(8km/h)": calories / 8,
            "快走(6km/h)": calories / 5,
            "游泳": calories / 10,
            "骑自行车": calories / 7,
            "瑜伽": calories / 3
        }
        result = "🏃 消耗这些热量需要运动：\n"
        for exercise, minutes in exercises.items():
            result += f"  • {exercise}: {minutes:.0f} 分钟\n"
        return result

    def generate_advice(self, intake, recommend):
        advice = "💡 个性化饮食建议：\n"
        goal = self.profile["goal"]
        disease = self.profile["disease"]

        calorie_ratio = intake["热量(kcal)"] / recommend["热量"] if recommend["热量"] > 0 else 1
        if calorie_ratio > 1.1:
            advice += "⚠️  今日热量摄入超标，建议减少精制碳水与油炸食品摄入\n"
        elif calorie_ratio < 0.9:
            advice += "⚠️  今日热量摄入不足，建议增加优质蛋白与复合碳水摄入\n"
        else:
            advice += "✅ 今日热量摄入符合目标，继续保持\n"

        if intake["蛋白质(g)"] < recommend["蛋白质"] * 0.8:
            advice += "⚠️  蛋白质摄入不足，建议增加鸡蛋、牛奶、鸡胸肉、鱼虾等优质蛋白\n"

        if intake["钠(mg)"] > recommend["钠"]:
            advice += "⚠️  钠摄入超标，建议减少咸菜、加工食品、高盐调味品摄入\n"

        if disease == "糖尿病":
            if intake["碳水(g)"] > recommend["碳水"] * 1.05:
                advice += "⚠️  碳水摄入超标，建议选择低GI食物，少食多餐\n"
        if disease == "高血压":
            advice += "⚠️  高血压患者需严格控制钠摄入，每日不超过2000mg\n"

        return advice

    def update_profile(self, **kwargs):
        self.profile.update(kwargs)
        self._save_profile()
        return self.get_profile_info()

    def get_profile_info(self):
        recommend = self.calculate_recommend_intake()
        p = self.profile
        info = f"👤 个人档案\n"
        info += f"姓名：{p['name']} | 性别：{p['gender']} | 年龄：{p['age']}岁\n"
        info += f"身高：{p['height']}cm | 体重：{p['weight']}kg\n"
        info += f"活动量：{p['activity']} | 目标：{p['goal']}\n"
        info += f"健康状况：{p['disease']}\n"
        info += "-" * 50 + "\n"
        info += f"📊 每日推荐摄入量：\n"
        info += f"热量：{recommend['热量']} kcal\n"
        info += f"蛋白质：{recommend['蛋白质']} g\n"
        info += f"碳水：{recommend['碳水']} g\n"
        info += f"脂肪：{recommend['脂肪']} g\n"
        info += f"钠：{recommend['钠']} mg\n"
        return info

    def _create_weekly_chart(self, df):
        """生成包含热量/蛋白质/碳水/脂肪四个维度的周趋势图"""
        daily = df.groupby('日期').agg({
            '热量(kcal)': 'sum', '蛋白质(g)': 'sum',
            '碳水(g)': 'sum', '脂肪(g)': 'sum'
        }).sort_index()

        # 补齐7天空缺日期
        today = datetime.date.today()
        all_dates = [(today - datetime.timedelta(days=i)).isoformat() for i in range(6, -1, -1)]
        daily = daily.reindex(all_dates, fill_value=0)

        dates = [d[-5:] for d in all_dates]  # 只显示 MM-DD
        metrics = [
            ('热量(kcal)', '热量 (kcal)', '#FF6B6B'),
            ('蛋白质(g)', '蛋白质 (g)', '#4ECDC4'),
            ('碳水(g)', '碳水 (g)', '#FFE66D'),
            ('脂肪(g)', '脂肪 (g)', '#95E1D3'),
        ]

        fig, axes = plt.subplots(2, 2, figsize=(12, 8))
        for ax, (col, label, color) in zip(axes.flat, metrics):
            values = daily[col].values
            ax.bar(dates, values, color=color, alpha=0.8)
            ax.set_title(label, fontsize=12)
            ax.set_xticklabels(dates, rotation=45, fontsize=8)
            ax.set_ylabel(label.split('(')[1].rstrip(')') if '(' in label else '')

        fig.suptitle('本周营养摄入趋势', fontsize=14, fontweight='bold')
        plt.tight_layout()

        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=100)
        buf.seek(0)
        plt.close()
        return Image.open(buf)

    def export_to_excel(self, date):
        """导出带格式和汇总的Excel膳食报告"""
        std_date = _standardize_date(date)
        daily_records = [r for r in self.records if r["日期"] == std_date]
        if not daily_records:
            return None

        wb = Workbook()
        ws = wb.active
        ws.title = f"膳食报告_{std_date}"

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        cell_align = Alignment(horizontal="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        warning_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        ok_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")

        # 表头
        headers = ["食品组合", "总重量(g)", "餐次", "热量(kcal)", "蛋白质(g)",
                    "碳水(g)", "脂肪(g)", "膳食纤维(g)", "钠(mg)", "平均GI值"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 数据行
        for row_idx, record in enumerate(daily_records, 2):
            for col_idx, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=record.get(key, ""))
                cell.alignment = cell_align
                cell.border = thin_border

        # 汇总行
        summary_row = len(daily_records) + 2
        summary_fill = PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")
        ws.cell(row=summary_row, column=1, value="📊 今日总计").font = Font(bold=True)

        numeric_cols = {"热量(kcal)": 4, "蛋白质(g)": 5, "碳水(g)": 6,
                         "脂肪(g)": 7, "膳食纤维(g)": 8, "钠(mg)": 9}
        for key, col_idx in numeric_cols.items():
            total_val = sum(r[key] for r in daily_records)
            cell = ws.cell(row=summary_row, column=col_idx, value=round(total_val, 1))
            cell.font = Font(bold=True)
            cell.fill = summary_fill
            cell.border = thin_border

        # 推荐摄入量对比行
        recommend = self.calculate_recommend_intake()
        rec_row = summary_row + 1
        ws.cell(row=rec_row, column=1, value="🎯 推荐摄入量").font = Font(bold=True, color="059669")
        rec_map = {4: recommend["热量"], 5: recommend["蛋白质"],
                    6: recommend["碳水"], 7: recommend["脂肪"], 9: recommend["钠"]}
        for col_idx, val in rec_map.items():
            cell = ws.cell(row=rec_row, column=col_idx, value=round(val, 1))
            cell.font = Font(bold=True, color="059669")
            cell.border = thin_border

        # 达标判断行
        judge_row = rec_row + 1
        ws.cell(row=judge_row, column=1, value="✅/⚠️ 达标判断").font = Font(bold=True)
        for key, col_idx in numeric_cols.items():
            actual = sum(r[key] for r in daily_records)
            target = rec_map.get(col_idx, 0)
            if target > 0:
                ratio = actual / target
                if 0.9 <= ratio <= 1.1:
                    status, fill = "✅ 达标", ok_fill
                elif ratio > 1.1:
                    status, fill = "⚠️ 超标", warning_fill
                else:
                    status, fill = "⚠️ 不足", warning_fill
            else:
                status, fill = "-", None
            cell = ws.cell(row=judge_row, column=col_idx, value=status)
            if fill:
                cell.fill = fill
            cell.border = thin_border

        # 嵌入餐次分布饼图
        meal_cal = {}
        for r in daily_records:
            meal = r.get("餐次", "午餐")
            meal_cal[meal] = meal_cal.get(meal, 0) + r.get("热量(kcal)", 0)

        if meal_cal:
            fig, ax = plt.subplots(figsize=(4, 3))
            colors_pie = ['#FF6B6B', '#4ECDC4', '#FFE66D', '#95E1D3']
            ax.pie(meal_cal.values(), labels=meal_cal.keys(), autopct='%1.1f%%',
                   colors=colors_pie[:len(meal_cal)], startangle=90)
            ax.set_title(f'{std_date} 各餐次热量分布')
            pie_buf = io.BytesIO()
            plt.savefig(pie_buf, format='png', dpi=100, bbox_inches='tight')
            pie_buf.seek(0)
            plt.close()

            img = XLImage(pie_buf)
            img.width, img.height = 300, 225
            ws.add_image(img, f"A{judge_row + 2}")

        # 列宽自适应
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14

        excel_file = os.path.join(self.user_dir, f"report_{std_date}.xlsx")
        wb.save(excel_file)
        return excel_file

    def _create_progress_rings(self, intake, recommend):
        """生成环形进度图，直观展示营养指标完成度"""
        metrics = [
            ("热量", intake["热量(kcal)"], recommend.get("热量", 2000), "kcal", "#FF6B6B"),
            ("蛋白质", intake["蛋白质(g)"], recommend.get("蛋白质", 60), "g", "#4ECDC4"),
            ("碳水", intake["碳水(g)"], recommend.get("碳水", 250), "g", "#FFE66D"),
            ("脂肪", intake["脂肪(g)"], recommend.get("脂肪", 60), "g", "#95E1D3"),
        ]

        fig, axes = plt.subplots(1, 4, figsize=(14, 3.5))

        for ax, (name, value, target, unit, color) in zip(axes, metrics):
            ratio = min(value / target, 1.0) if target > 0 else 0
            ring_color = "#059669" if ratio <= 1.0 else "#DC2626"

            # 绘制环形进度
            theta = np.linspace(0, 2 * np.pi, 100)
            r_inner, r_outer = 0.6, 1.0

            # 背景圆环
            ax.fill_between(np.cos(theta) * r_outer, np.cos(theta) * r_inner,
                           np.sin(theta) * r_outer, np.sin(theta) * r_inner,
                           color='#E5E7EB', alpha=0.5)

            # 进度弧
            if ratio > 0:
                progress_theta = np.linspace(0, 2 * np.pi * ratio, max(3, int(100 * ratio)))
                ax.fill_between(np.cos(progress_theta) * r_outer, np.cos(progress_theta) * r_inner,
                               np.sin(progress_theta) * r_outer, np.sin(progress_theta) * r_inner,
                               color=ring_color, alpha=0.9)

            # 中心文字
            ax.text(0, 0.15, f"{value:.0f}", ha='center', va='center', fontsize=16, fontweight='bold')
            ax.text(0, -0.2, f"/ {target:.0f} {unit}", ha='center', va='center', fontsize=9, color='#6B7280')
            ax.text(0, -0.5, f"{ratio*100:.0f}%", ha='center', va='center', fontsize=10,
                   color=ring_color, fontweight='bold')

            ax.set_xlim(-1.3, 1.3)
            ax.set_ylim(-1.3, 1.3)
            ax.set_aspect('equal')
            ax.axis('off')
            ax.set_title(name, fontsize=13, fontweight='bold', pad=8)

        plt.tight_layout()

        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=120, bbox_inches='tight', transparent=True)
        buf.seek(0)
        plt.close()
        return Image.open(buf)

    def generate_daily_summary(self):
        """生成AI助手每日开场白，基于规则引擎，不依赖大模型"""
        today = datetime.date.today().isoformat()
        daily_records = [r for r in self.records if r["日期"] == today]
        recommend = self.calculate_recommend_intake()

        if not daily_records:
            return f"📋 主任医师今日查房（{today}）：\n\n今日暂无进食记录，记得按时用餐哦！\n建议三餐规律饮食，保证营养均衡。\n\n请问今天有什么可以帮到您的？"

        # 汇总今日摄入
        total = {"热量(kcal)": 0, "蛋白质(g)": 0, "碳水(g)": 0,
                  "脂肪(g)": 0, "膳食纤维(g)": 0, "钠(mg)": 0}
        for r in daily_records:
            for k in total:
                total[k] += r[k]

        summary = f"📋 主任医师今日查房小结（{today}）：\n\n"
        summary += "截至目前，您今日已摄入：\n"
        summary += f"• 热量：{total['热量(kcal)']:.0f} kcal（完成目标的 {total['热量(kcal)']/recommend['热量']*100:.0f}%）\n"
        summary += f"• 蛋白质：{total['蛋白质(g)']:.0f}g（完成目标的 {total['蛋白质(g)']/recommend['蛋白质']*100:.0f}%）\n"
        summary += f"• 碳水：{total['碳水(g)']:.0f}g / 推荐 {recommend['碳水']:.0f}g\n"
        summary += f"• 脂肪：{total['脂肪(g)']:.0f}g / 推荐 {recommend['脂肪']:.0f}g\n"

        # 风险提醒
        warnings = []
        if total["热量(kcal)"] > recommend["热量"] * 0.9:
            warnings.append("今日热量已接近全天推荐上限，晚餐建议清淡少食")
        if total["钠(mg)"] > recommend["钠"] * 0.7:
            warnings.append(f"钠摄入已达 {total['钠(mg)']:.0f}mg，请注意控盐")
        if total["蛋白质(g)"] < recommend["蛋白质"] * 0.4:
            warnings.append("蛋白质摄入偏少，晚餐建议补充优质蛋白（鱼/鸡胸肉/豆腐）")
        if total["膳食纤维(g)"] < 10:
            warnings.append("膳食纤维摄入不足，建议增加蔬菜摄入")

        # 检查是否有水果摄入
        has_fruit = any(
            any(fruit in r.get("食品组合", "") for fruit in ["苹果", "香蕉", "葡萄", "橙子", "菠萝", "西瓜"])
            for r in daily_records
        )
        if not has_fruit:
            warnings.append("今日尚未摄入水果，建议补充一份低GI水果（如苹果、橙子）")

        if warnings:
            summary += "\n⚠️ 需要关注：\n"
            for w in warnings:
                summary += f"• {w}\n"
        else:
            summary += "\n✅ 今日饮食结构良好，继续保持！\n"

        summary += "\n请问今天有什么可以帮到您的？"
        return summary


# -------------------------- 4. 核心功能函数（多用户版）--------------------------

def _standardize_date(date_str):
    """将各种格式的日期字符串统一转换为 YYYY-MM-DD"""
    if not date_str:
        return datetime.date.today().isoformat()
    try:
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
            try:
                dt = datetime.datetime.strptime(str(date_str).strip(), fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue
        return str(date_str).strip()
    except Exception:
        return datetime.date.today().isoformat()


def predict_food(image):
    """图像识别食物（不依赖用户状态）"""
    if image is None:
        return "请上传食物图片！", [], None, None, 0

    try:
        results = model.predict(
            source=image,
            conf=0.25,
            imgsz=640,
            device=device,
            verbose=False
        )

        boxes = results[0].boxes
        if len(boxes) == 0:
            return "未识别到食物，请更换清晰图片！", [], None, None, 0

        total_nutrition = [0] * 7
        detected_items = {}
        estimated_total_weight = 0
        annotated_image = results[0].plot()
        annotated_pil = Image.fromarray(annotated_image[..., ::-1])

        for box in boxes:
            class_id = int(box.cls)
            if class_id < len(FOOD_CLASSES):
                food_name = FOOD_CLASSES[class_id]
                detected_items[food_name] = detected_items.get(food_name, 0) + 1

        # 基于默认份量估算每种食物的实际营养
        for food_name, count in detected_items.items():
            portion_w = DEFAULT_PORTION_WEIGHT.get(food_name, 100)
            food_weight = portion_w * count
            estimated_total_weight += food_weight
            nutr = NUTRITION_DB.get(food_name, [0] * 7)
            for i in range(7):
                total_nutrition[i] += nutr[i] * (food_weight / 100)

        if estimated_total_weight > 0:
            nutrition_per_100g = [v * 100 / estimated_total_weight for v in total_nutrition]
        else:
            nutrition_per_100g = total_nutrition

        summary_parts = [f"{k}x{v}" for k, v in detected_items.items()]
        summary_str = ", ".join(summary_parts)

        return summary_str, nutrition_per_100g, annotated_pil, detected_items, estimated_total_weight
    except FileNotFoundError as e:
        logger.error(f"模型文件未找到: {e}")
        return f"❌ 模型文件缺失：{str(e)}\n请确认 YOLO 模型路径是否正确", [], None, None, 0
    except RuntimeError as e:
        logger.error(f"模型推理错误: {e}")
        return f"❌ 模型推理失败（可能是显存不足或图片格式问题）：{str(e)}", [], None, None, 0
    except Exception as e:
        logger.error(f"识别异常: {e}", exc_info=True)
        return f"识别出错：{str(e)}", [], None, None, 0


def main_predict(username, image, weight, date, meal_type):
    """图像识别并记录 —— 多用户版"""
    if not username:
        return "❌ 请先登录", "", None

    dm = DietManager.get_instance(username)
    summary, total_nutrition_per_100g, annotated_img, items_dict, estimated_weight = predict_food(image)

    if not items_dict:
        return f"识别结果：{summary}\n❌ 无法添加记录", "", annotated_img

    actual_weight = weight if weight != 100 else estimated_weight
    weight_source = "（系统估算）" if weight == 100 and estimated_weight > 0 else "（用户输入）"

    nutrition_info = f"🍽️  识别结果：{summary}\n"
    nutrition_info += "=" * 40 + "\n"
    nutrition_info += f"📊 系统估算总重量：{estimated_weight}g（默认份量 × 检测数量）\n"
    nutrition_info += f"📊 综合营养成分（每100g混合平均值）：\n"
    nutrition_info += f"热量：{total_nutrition_per_100g[0]:.1f} kcal\n"
    nutrition_info += f"蛋白质：{total_nutrition_per_100g[1]:.1f} g\n"
    nutrition_info += f"碳水：{total_nutrition_per_100g[2]:.1f} g\n"
    nutrition_info += f"脂肪：{total_nutrition_per_100g[3]:.1f} g\n"
    nutrition_info += f"膳食纤维：{total_nutrition_per_100g[4]:.1f} g\n"
    nutrition_info += f"钠：{total_nutrition_per_100g[5]:.1f} mg\n"

    nutrition_info += "=" * 40 + "\n"
    nutrition_info += f"🍴 本次食用总重 {actual_weight}g {weight_source}预计摄入：\n"
    ratio = actual_weight / 100
    nutrition_info += f"总热量：{total_nutrition_per_100g[0] * ratio:.1f} kcal\n"
    nutrition_info += f"总蛋白质：{total_nutrition_per_100g[1] * ratio:.1f} g\n"
    nutrition_info += f"总碳水：{total_nutrition_per_100g[2] * ratio:.1f} g\n"

    if weight == 100 and estimated_weight > 0:
        nutrition_info += f"\n💡 提示：以上为系统估算值，如实际称重不同，请修改\"总食用重量\"后重新识别\n"

    std_date = _standardize_date(date)
    record_msg = dm.add_record(summary, actual_weight, total_nutrition_per_100g, std_date, meal_type)

    return nutrition_info, record_msg, annotated_img


def query_records(username, date):
    """查询记录 —— 多用户版"""
    if not username:
        return "❌ 请先登录", None, None
    dm = DietManager.get_instance(username)
    std_date = _standardize_date(date)
    report, df, _, progress_img = dm.get_daily_stats(std_date)
    return report, df, progress_img


def delete_record_handler(username, date, index):
    """删除指定日期的某条膳食记录 —— 多用户版"""
    if not username:
        return "❌ 请先登录", "", None, None
    dm = DietManager.get_instance(username)
    std_date = _standardize_date(date)
    indices = [i for i, r in enumerate(dm.records) if r["日期"] == std_date]
    if not indices:
        return "❌ 当日无记录可删除", "", None, None
    if 0 <= index < len(indices):
        result = dm.delete_record(indices[index])
        report, df, _, progress_img = dm.get_daily_stats(std_date)
        return result, report, df, progress_img
    return f"❌ 序号无效：有效范围 0~{len(indices)-1}，请检查", "", None, None


def manual_add_food(username, selected_foods, weights_str, date, meal_type):
    """手动输入食物记录 —— 多用户版"""
    if not username:
        return "❌ 请先登录", ""
    dm = DietManager.get_instance(username)

    if not selected_foods:
        return "❌ 请至少选择一种食物", ""

    food_names = list(selected_foods)

    unknown_foods = [f for f in food_names if f not in NUTRITION_DB]
    known_foods = [f for f in food_names if f in NUTRITION_DB]
    if unknown_foods:
        unknown_hint = "、".join(unknown_foods)
        if not known_foods:
            return f"❌ 以下食物不在营养数据库中：{unknown_hint}\n请从下拉列表中选择已有食物", ""

    weights = []
    if weights_str and weights_str.strip():
        try:
            weights = [float(w.strip()) for w in weights_str.split(",") if w.strip()]
        except ValueError:
            return "❌ 重量格式错误，请用逗号分隔数字", ""

    if len(weights) != len(food_names):
        weights = [DEFAULT_PORTION_WEIGHT.get(f, 100) for f in food_names]

    total_weight = sum(weights)
    total_nutrition = [0] * 7
    for food_name, w in zip(food_names, weights):
        nutr = NUTRITION_DB.get(food_name, [0] * 7)
        for i in range(7):
            total_nutrition[i] += nutr[i] * (w / 100)

    if total_weight > 0:
        nutrition_per_100g = [v * 100 / total_weight for v in total_nutrition]
    else:
        nutrition_per_100g = total_nutrition

    summary = ", ".join(f"{f}x1" for f in food_names)
    std_date = _standardize_date(date)
    record_msg = dm.add_record(summary, total_weight, nutrition_per_100g, std_date, meal_type)

    info = f"🍽️ 手动录入：{summary}\n"
    info += f"总重量：{total_weight}g\n"
    info += f"热量：{nutrition_per_100g[0] * total_weight / 100:.1f} kcal\n"
    info += f"蛋白质：{nutrition_per_100g[1] * total_weight / 100:.1f} g\n"
    info += f"碳水：{nutrition_per_100g[2] * total_weight / 100:.1f} g\n"
    info += f"脂肪：{nutrition_per_100g[3] * total_weight / 100:.1f} g\n"
    if unknown_foods:
        info += f"\n⚠️ 已跳过未知食物：{unknown_hint}"

    return info, record_msg


def get_daily_summary_for_ai(username):
    """获取AI助手每日开场白 —— 多用户版"""
    if not username:
        return "❌ 请先登录"
    dm = DietManager.get_instance(username)
    return dm.generate_daily_summary()


def create_dashboard(username):
    """生成今日概览仪表盘 —— 多用户版"""
    if not username:
        return "❌ 请先登录", None
    dm = DietManager.get_instance(username)
    today = datetime.date.today().isoformat()
    _, df, intake, progress_img = dm.get_daily_stats(today)
    summary = dm.generate_daily_summary()

    if df is not None and not df.empty:
        recent = df.tail(3)[['餐次', '食品组合', '热量(kcal)']].to_string(index=False)
    else:
        recent = "今日暂无记录"

    dash_text = f"📊 今日概览（{today}）\n"
    dash_text += "=" * 40 + "\n"
    dash_text += f"📋 最近记录：\n{recent}\n"
    dash_text += "-" * 40 + "\n"
    dash_text += summary.split("请问今天有什么可以帮到您的？")[0]

    return dash_text, progress_img


def update_profile_handler(username, name, height, weight, age, gender, activity, goal, disease):
    """更新个人档案 —— 多用户版"""
    if not username:
        return "❌ 请先登录"
    dm = DietManager.get_instance(username)
    return dm.update_profile(
        name=name, height=height, weight=weight, age=age,
        gender=gender, activity=activity, goal=goal, disease=disease
    )


def update_chat_avatar(username):
    """
    根据用户性别返回 gr.update 用于动态更新 chat_bot 头像
    - 男 → boy.jpg，女 → girl.jpg
    - 医生头像固定为 doctor.jpg
    - 文件缺失时静默降级（尝试另一性别，都缺失则留空）
    """
    doctor_path = os.path.join(PICTURE_DIR, "doctor.jpg")
    if not os.path.exists(doctor_path):
        doctor_path = ""

    # 默认男头像
    default_user_path = os.path.join(PICTURE_DIR, "boy.jpg")
    if not username:
        return gr.update(avatar_images=(default_user_path if os.path.exists(default_user_path) else "", doctor_path))

    # 读取用户性别
    try:
        dm = DietManager.get_instance(username)
        gender = dm.profile.get("gender", "男")
    except Exception:
        gender = "男"

    selected = "boy.jpg" if gender == "男" else "girl.jpg"
    user_path = os.path.join(PICTURE_DIR, selected)

    # 降级逻辑：目标文件不存在时尝试另一性别头像
    if not os.path.exists(user_path):
        fallback = "girl.jpg" if gender == "男" else "boy.jpg"
        fallback_path = os.path.join(PICTURE_DIR, fallback)
        if os.path.exists(fallback_path):
            user_path = fallback_path
        else:
            user_path = ""

    return gr.update(avatar_images=(user_path, doctor_path))


def load_profile_to_form(username):
    """
    根据用户名加载个人档案，返回所有档案输入框的 gr.update 对象
    - 登录后自动填充已保存数据
    - 保存后刷新输入框
    - 退出登录时重置为默认值
    """
    if not username:
        # 未登录 → 默认值
        return (
            gr.update(value="用户"),
            gr.update(value=170),
            gr.update(value=65),
            gr.update(value=20),
            gr.update(value="男"),
            gr.update(value="轻度活动"),
            gr.update(value="维持体重"),
            gr.update(value="健康")
        )

    try:
        dm = DietManager.get_instance(username)
        if dm is None:
            return load_profile_to_form("")
        p = dm.profile
        return (
            gr.update(value=p.get("name", username)),
            gr.update(value=p.get("height", 170)),
            gr.update(value=p.get("weight", 65)),
            gr.update(value=p.get("age", 20)),
            gr.update(value=p.get("gender", "男")),
            gr.update(value=p.get("activity", "轻度活动")),
            gr.update(value=p.get("goal", "维持体重")),
            gr.update(value=p.get("disease", "健康"))
        )
    except Exception as e:
        logger.error(f"加载档案到表单失败 ({username}): {e}")
        return load_profile_to_form("")


def show_weekly_trend(username):
    """展示周趋势 —— 多用户版"""
    if not username:
        return "❌ 请先登录", None
    dm = DietManager.get_instance(username)
    msg, chart = dm.get_weekly_stats()
    return msg, chart


def export_report(username, date):
    """导出Excel报告 —— 多用户版"""
    if not username:
        return "❌ 请先登录"
    dm = DietManager.get_instance(username)
    std_date = _standardize_date(date)
    try:
        file_path = dm.export_to_excel(std_date)
        if file_path:
            return f"✅ 报告已成功导出至项目文件夹：{file_path}"
        return "❌ 当日无记录，无法导出。"
    except ImportError:
        return "❌ 导出失败：缺少底层依赖库。请在终端运行 `pip install openpyxl`"
    except PermissionError:
        return "❌ 导出失败：该 Excel 文件正在被其他程序（如 Office/WPS）占用，请先关闭文件后再试！"
    except Exception as e:
        return f"❌ 导出遇到未知错误：{str(e)}"


def load_profile_for_user(username):
    """登录后加载用户档案信息 —— 多用户版"""
    if not username:
        return "❌ 请先登录"
    dm = DietManager.get_instance(username)
    return dm.get_profile_info()


# -------------------------- 4.1 聊天历史管理（多用户持久化）--------------------------

def _get_chat_history_path(username):
    """获取用户聊天历史文件的存储路径"""
    return os.path.join(DATA_DIR, username, "chat_history.json")


def load_user_chat_history(username):
    """
    加载用户的完整聊天历史数据
    若文件不存在或损坏则自动重建默认结构，确保程序不崩溃
    """
    filepath = _get_chat_history_path(username)
    if not os.path.exists(filepath):
        return _create_default_chat_history(username)

    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        # 校验数据结构完整性
        if "sessions" not in data or not isinstance(data["sessions"], list):
            raise ValueError("数据结构异常：缺少 sessions 字段")
        if not data["sessions"]:
            # 空会话列表是合法的（用户可能删除了所有会话），不自动重建
            data["current_session_id"] = ""
            return data
        if not data.get("current_session_id") or not any(
            s["session_id"] == data["current_session_id"] for s in data["sessions"]
        ):
            data["current_session_id"] = data["sessions"][0]["session_id"]
            _save_chat_history_raw(username, data)
        return data
    except (json.JSONDecodeError, ValueError, IOError) as e:
        logger.warning(f"聊天历史文件损坏，自动重建 ({username}): {e}")
        # 备份损坏文件以便排查
        if os.path.exists(filepath):
            try:
                shutil.copy2(filepath, filepath + ".corrupted")
            except Exception:
                pass
        return _create_default_chat_history(username)


def _create_default_chat_history(username):
    """创建默认的空聊天历史结构（含一个初始空会话）"""
    session_id = _generate_chat_session_id()
    now = datetime.datetime.now().isoformat()
    data = {
        "sessions": [{
            "session_id": session_id,
            "title": "新对话",
            "created_at": now,
            "updated_at": now,
            "messages": []
        }],
        "current_session_id": session_id
    }
    _save_chat_history_raw(username, data)
    return data


def _save_chat_history_raw(username, data):
    """直接写入聊天历史到磁盘（内部方法）"""
    filepath = _get_chat_history_path(username)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    if os.path.exists(filepath):
        try:
            shutil.copy2(filepath, filepath + ".bak")
        except Exception:
            pass
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def save_user_chat_history(username, data):
    """保存用户聊天历史（对外接口，带校验）"""
    if not username or not data:
        return
    _save_chat_history_raw(username, data)


def _generate_chat_session_id():
    """生成唯一的会话ID：时间戳 + UUID短码"""
    return f"{int(datetime.datetime.now().timestamp() * 1000)}_{uuid.uuid4().hex[:8]}"


def create_new_session(username):
    """
    创建新的聊天会话，自动设为当前会话
    返回新会话的 session_id
    """
    data = load_user_chat_history(username)
    session_id = _generate_chat_session_id()
    now = datetime.datetime.now().isoformat()
    new_session = {
        "session_id": session_id,
        "title": f"新对话 {datetime.datetime.now().strftime('%m-%d %H:%M')}",
        "created_at": now,
        "updated_at": now,
        "messages": []
    }
    data["sessions"].insert(0, new_session)  # 最新会话置顶
    data["current_session_id"] = session_id
    _save_chat_history_raw(username, data)
    logger.info(f"用户 {username} 创建新聊天会话：{session_id}")
    return session_id


def switch_session(username, session_id):
    """
    切换到指定会话，返回 (chatbot消息列表, 会话标题, session_id)
    若会话不存在则自动回退到第一个可用会话
    """
    data = load_user_chat_history(username)
    session = None
    for s in data["sessions"]:
        if s["session_id"] == session_id:
            session = s
            break

    if session is None:
        # 回退到第一个会话
        if data["sessions"]:
            session = data["sessions"][0]
            session_id = session["session_id"]
        else:
            session_id = create_new_session(username)
            data = load_user_chat_history(username)
            session = data["sessions"][0]

    data["current_session_id"] = session_id
    _save_chat_history_raw(username, data)

    # 转换为 Gradio Chatbot 格式（dict 列表）
    chatbot_msgs = []
    for msg in session.get("messages", []):
        chatbot_msgs.append({"role": msg["role"], "content": msg["content"]})

    return chatbot_msgs, session.get("title", "新对话"), session_id


def get_sessions_for_display(username):
    """
    获取会话列表的 DataFrame（用于侧边栏展示）以及对应的 session_id 列表
    返回 (pd.DataFrame, list_of_session_ids)
    """
    data = load_user_chat_history(username)
    rows = []
    ids = []
    for s in data["sessions"]:
        title = s.get("title", "新对话")
        display_title = title[:25] + "..." if len(title) > 25 else title
        updated = s.get("updated_at", s.get("created_at", ""))
        try:
            dt = datetime.datetime.fromisoformat(updated)
            time_str = dt.strftime("%m-%d %H:%M")
        except Exception:
            time_str = updated[:16] if len(updated) >= 16 else updated
        rows.append({"💬 会话": display_title, "🕒 更新时间": time_str})
        ids.append(s["session_id"])

    if not rows:
        rows.append({"💬 会话": "暂无会话", "🕒 更新时间": ""})
        ids.append("")
    return pd.DataFrame(rows), ids


def persist_chat_message(username, session_id, user_msg, assistant_msg):
    """将一轮完整的对话（用户消息+AI回复）持久化写入 JSON 文件"""
    data = load_user_chat_history(username)
    for s in data["sessions"]:
        if s["session_id"] == session_id:
            s["messages"].append({"role": "user", "content": user_msg})
            s["messages"].append({"role": "assistant", "content": assistant_msg})
            s["updated_at"] = datetime.datetime.now().isoformat()
            # 首个用户消息自动成为会话标题
            if len(s["messages"]) == 2:
                title = user_msg[:20] + ("..." if len(user_msg) > 20 else "")
                s["title"] = title
            break
    _save_chat_history_raw(username, data)


def delete_chat_session(username, session_id):
    """
    删除指定会话，处理所有边界情况
    返回 (new_current_session_id_or_None, deleted_title)
    - 若删除后仍有会话且删除的是当前会话 → 自动切换到第一个剩余会话
    - 若删除的是非当前会话 → current_session_id 保持不变
    - 若删除后无剩余会话 → 返回 (None, deleted_title)
    """
    data = load_user_chat_history(username)
    deleted_title = "未知会话"

    # 记录被删除会话的标题（用于反馈）
    for s in data["sessions"]:
        if s["session_id"] == session_id:
            deleted_title = s.get("title", "未知会话")
            break

    was_current = (data["current_session_id"] == session_id)

    # 执行删除
    data["sessions"] = [s for s in data["sessions"] if s["session_id"] != session_id]

    # 确定新的 current_session_id
    if data["sessions"]:
        if was_current:
            data["current_session_id"] = data["sessions"][0]["session_id"]
        # 非当前会话被删除，current_session_id 保持不变
    else:
        data["current_session_id"] = ""

    _save_chat_history_raw(username, data)
    logger.info(f"用户 {username} 删除会话「{deleted_title}」（{session_id}），剩余 {len(data['sessions'])} 个会话")

    return (data["current_session_id"] if data["sessions"] else None), deleted_title


def _make_session_title_html(title):
    """生成会话标题的 HTML 片段（右侧聊天区顶部）"""
    return f"""
    <div class="fm-session-title">
        <span>💬 {title}</span>
    </div>
    """


def init_chat_state(username):
    """
    用户登录/切换时初始化聊天状态
    返回 (session_id, sessions_df, session_ids_list, title_html, chatbot_msgs)
    """
    try:
        if not username:
            empty_df = pd.DataFrame({"💬 会话": [], "🕒 更新时间": []})
            return "", empty_df, [], "", []

        logger.info(f"初始化聊天状态：用户 {username}")
        data = load_user_chat_history(username)
        session_id = data.get("current_session_id", "")
        session = None
        for s in data["sessions"]:
            if s["session_id"] == session_id:
                session = s
                break
        if session is None and data["sessions"]:
            session = data["sessions"][0]
            session_id = session["session_id"]

        df, ids = get_sessions_for_display(username)

        if session:
            chatbot_msgs = [{"role": m["role"], "content": m["content"]} for m in session.get("messages", [])]
            title = session.get("title", "新对话")
            logger.info(f"聊天状态初始化完成：会话「{title}」，{len(chatbot_msgs)} 条消息，共 {len(data['sessions'])} 个会话")
            return session_id, df, ids, _make_session_title_html(title), chatbot_msgs
        else:
            new_id = create_new_session(username)
            df2, ids2 = get_sessions_for_display(username)
            logger.info(f"聊天状态初始化完成：自动创建新会话 {new_id}")
            return new_id, df2, ids2, _make_session_title_html("新对话"), []
    except Exception as e:
        logger.error(f"初始化聊天状态失败 ({username}): {e}", exc_info=True)
        empty_df = pd.DataFrame({"💬 会话": [], "🕒 更新时间": []})
        return "", empty_df, [], _make_session_title_html("加载失败，请刷新"), []


# -------------------------- 5. Ollama 服务检查 --------------------------

def check_ollama_service():
    """检查 Ollama 服务是否可用，返回 (是否可用, 提示信息)"""
    try:
        ollama_client.list()
        return True, ""
    except Exception as e:
        error_msg = str(e)
        if "ConnectionError" in type(e).__name__ or "connection" in error_msg.lower():
            return False, "❌ 无法连接到 Ollama 服务。\n\n请确保：\n1. 已安装 Ollama（https://ollama.com）\n2. Ollama 服务正在运行（终端执行 `ollama serve`）\n3. 模型已下载（`ollama pull qwen2.5:7b`）"
        else:
            logger.warning(f"Ollama 服务检查异常: {error_msg}")
            return False, f"❌ Ollama 服务异常：{error_msg}"


# -------------------------- 6. AI 聊天引擎（多用户版）--------------------------

def chat_with_ai(user_message, history, username):
    """AI 聊天 —— 多用户版"""
    try:
        if not username:
            yield "❌ 请先登录后再使用AI咨询功能"
            return

        dm = DietManager.get_instance(username)

        # Ollama 服务可用性检查
        ollama_ok, ollama_err = check_ollama_service()
        if not ollama_ok:
            yield ollama_err
            return

        # 获取用户上下文
        profile_context = dm.get_profile_info()
        today_date = datetime.date.today().isoformat()
        _, today_records_df, _, _ = dm.get_daily_stats(today_date)

        records_str = "今日暂无进食记录"
        if today_records_df is not None and not today_records_df.empty:
            records_str = today_records_df[['食品组合', '总食用重量(g)', '餐次', '热量(kcal)']].to_string(index=False)

        system_prompt = f"""你是一名著名三甲医院临床营养科的主任医师，现在作为用户的专属膳食管家进行问诊。
请严格基于以下用户的真实体征档案和今日饮食记录，提供专业、严谨、客观的医学解答。

【用户健康档案】：
{profile_context}

【今日膳食摄入记录】：
{records_str}

你的回复必须严格遵守以下临床规范：
1. 语气与口吻：严肃、专业、客观。采用门诊医生的沟通风格，直切主题，不讲废话。
2. 专业术语：必须使用标准的医学和营养学名词。
3. 数据驱动：所有的建议必须基于用户档案中的具体数据进行量化分析。
4. 风险警示：发现饮食记录存在超标风险，必须给出直接、严肃的临床医学警告。
5. 边界感：明确你的身份是营养医师，不进行其他疾病诊断。
"""

        messages = [{"role": "system", "content": system_prompt}]

        # 解析历史记录
        for item in history:
            if isinstance(item, dict):
                role = item.get("role", "user")
                raw_content = item.get("content", "")

                clean_text = ""
                if isinstance(raw_content, str):
                    clean_text = raw_content
                elif isinstance(raw_content, list):
                    for part in raw_content:
                        if isinstance(part, dict) and "text" in part:
                            clean_text += part["text"]
                        elif isinstance(part, str):
                            clean_text += part

                if clean_text.strip():
                    messages.append({"role": role, "content": clean_text})

            elif isinstance(item, (list, tuple)) and len(item) >= 2:
                human_text, assistant_text = item[0], item[1]
                if human_text:
                    messages.append({"role": "user", "content": str(human_text)})
                if assistant_text:
                    messages.append({"role": "assistant", "content": str(assistant_text)})

        messages.append({"role": "user", "content": str(user_message)})

        partial_message = ""

        response = ollama_client.chat(
            model='qwen2.5:7b',
            messages=messages,
            stream=True,
            options={'num_ctx': 8192}
        )

        for chunk in response:
            content = chunk.get('message', {}).get('content', '')
            if content:
                partial_message += content
                yield partial_message

    except ConnectionError as e:
        logger.error(f"Ollama 连接错误: {e}")
        yield f"⚠️ 与本地AI服务的连接中断。\n\n💡 请检查 Ollama 是否仍在运行（终端执行 `ollama serve` 重启服务）"
    except Exception as e:
        logger.error(f"AI 聊天异常: {e}", exc_info=True)
        yield f"⚠️ 抱歉，诊室网络出现波动。\n底层报错详情：{str(e)}\n\n💡 提示：请查看运行日志 app.log 了解详情。"


# -------------------------- 7. 登录/注册界面处理函数 --------------------------

def handle_login(username, password):
    """处理登录请求，返回 (username_state, login_block_visible, main_block_visible, message, ...)"""
    if not username or not username.strip():
        return "", gr.update(visible=True), gr.update(visible=False), gr.update(visible=True), gr.update(visible=False), "❌ 请输入用户名"
    if not password:
        return "", gr.update(visible=True), gr.update(visible=False), gr.update(visible=True), gr.update(visible=False), "❌ 请输入密码"

    username = username.strip()
    success, message = verify_login(username, password)
    if success:
        # 登录成功：隐藏登录区块，显示主应用区块
        return (
            username,                                    # current_user state
            gr.update(visible=False),                    # login_block 隐藏
            gr.update(visible=True),                     # main_block 显示
            gr.update(visible=True),                     # login_form 显示（重置）
            gr.update(visible=False),                    # register_form 隐藏（重置）
            message                                      # auth_message
        )
    else:
        return (
            "",                                          # current_user 保持空
            gr.update(visible=True),                     # login_block 保持显示
            gr.update(visible=False),                    # main_block 保持隐藏
            gr.update(visible=True),                     # login_form 保持显示
            gr.update(visible=False),                    # register_form 保持隐藏
            message                                      # auth_message
        )


def handle_register(username, password, confirm_password):
    """处理注册请求"""
    if not username or not username.strip():
        return "", gr.update(visible=True), gr.update(visible=False), gr.update(visible=False), gr.update(visible=True), "❌ 请输入用户名"
    if not password:
        return "", gr.update(visible=True), gr.update(visible=False), gr.update(visible=False), gr.update(visible=True), "❌ 请输入密码"
    if password != confirm_password:
        return "", gr.update(visible=True), gr.update(visible=False), gr.update(visible=False), gr.update(visible=True), "❌ 两次输入的密码不一致"

    username = username.strip()
    success, message = register_user(username, password)
    if success:
        # 注册成功直接登录
        return (
            username,                                    # current_user state
            gr.update(visible=False),                    # login_block 隐藏
            gr.update(visible=True),                     # main_block 显示
            gr.update(visible=False),                    # login_form 隐藏
            gr.update(visible=True),                     # register_form 显示（重置用）
            message                                      # auth_message
        )
    else:
        return (
            "",                                          # current_user 保持空
            gr.update(visible=True),                     # login_block 保持显示
            gr.update(visible=False),                    # main_block 保持隐藏
            gr.update(visible=False),                    # login_form 保持隐藏
            gr.update(visible=True),                     # register_form 保持显示
            message                                      # auth_message
        )


def handle_logout():
    """处理退出登录，返回所有需要重置的状态"""
    return (
        "",                             # current_user 清空
        gr.update(visible=True),        # login_block 显示
        gr.update(visible=False),       # main_block 隐藏
        gr.update(visible=True),        # login_form 显示
        gr.update(visible=False),       # register_form 隐藏
        "已安全退出登录"                # auth_message
    )


def switch_to_register():
    """切换到注册表单"""
    return (
        gr.update(visible=False),   # login_form 隐藏
        gr.update(visible=True),    # register_form 显示
        ""                          # 清空提示消息
    )


def switch_to_login():
    """切换到登录表单"""
    return (
        gr.update(visible=True),    # login_form 显示
        gr.update(visible=False),   # register_form 隐藏
        ""                          # 清空提示消息
    )


def get_user_bar_html(username):
    """生成顶部用户信息栏的HTML"""
    if not username:
        return ""
    avatar_char = username.strip()[:1].upper() if username.strip() else "U"
    return f"""
    <div class="fm-user-chip">
        <span class="fm-user-avatar">{avatar_char}</span>
        <span class="fm-user-name">{username}</span>
        <span class="fm-user-dot" title="在线"></span>
    </div>
    """


# -------------------------- 7.1 聊天界面事件处理函数 --------------------------

def handle_new_chat(username):
    """
    新建聊天会话：创建新会话 → 清空聊天区 → 刷新侧边栏
    返回 (session_id, chatbot_msgs, session_title_html, session_df, session_ids)
    """
    try:
        if not username:
            return "", [], _make_session_title_html("新对话"), pd.DataFrame({"💬 会话": [], "🕒 更新时间": []}), []
        new_id = create_new_session(username)
        df, ids = get_sessions_for_display(username)
        logger.info(f"用户 {username} 新建聊天会话：{new_id}")
        return new_id, [], _make_session_title_html("新对话"), df, ids
    except Exception as e:
        logger.error(f"新建聊天失败 ({username}): {e}", exc_info=True)
        return "", [], _make_session_title_html("操作失败"), pd.DataFrame({"💬 会话": [], "🕒 更新时间": []}), []


def handle_session_select(username, session_ids, select_data: gr.SelectData):
    """
    用户点击侧边栏历史会话条目 → 切换会话并加载消息历史
    返回 (session_id, chatbot_msgs, session_title_html, msg_input_clear, session_df, session_ids)
    """
    try:
        if not username or select_data is None or not select_data.index:
            return gr.skip(), gr.skip(), gr.skip(), gr.skip(), gr.skip(), gr.skip()

        idx = select_data.index[0]
        if idx < 0 or idx >= len(session_ids) or not session_ids[idx]:
            return gr.skip(), gr.skip(), gr.skip(), gr.skip(), gr.skip(), gr.skip()

        session_id = session_ids[idx]
        msgs, title, sid = switch_session(username, session_id)
        df, ids = get_sessions_for_display(username)
        logger.info(f"用户 {username} 切换到会话「{title}」（{session_id}），{len(msgs)} 条消息")
        return sid, msgs, _make_session_title_html(title), "", df, ids
    except Exception as e:
        logger.error(f"切换会话失败 ({username}): {e}", exc_info=True)
        return gr.skip(), gr.skip(), gr.skip(), gr.skip(), gr.skip(), gr.skip()


def handle_send_message(username, session_id, user_message, chat_history):
    """
    流式发送消息 → 调用 chat_with_ai → 实时更新 Chatbot → 持久化到JSON
    这是生成器函数，Gradio 自动处理流式输出
    """
    # 未登录或空消息
    if not username:
        yield chat_history, session_id, "", gr.skip(), gr.skip()
        return
    if not user_message or not user_message.strip():
        yield chat_history, session_id, "", gr.skip(), gr.skip()
        return

    user_message = user_message.strip()

    # 确保有有效会话
    if not session_id:
        data = load_user_chat_history(username)
        session_id = data.get("current_session_id", "")
        if not session_id:
            session_id = create_new_session(username)

    # 追加用户消息到显示
    chat_history.append({"role": "user", "content": user_message})

    # 构建传给 chat_with_ai 的历史（当前用户消息之前的全部对话）
    history_for_ai = chat_history[:-1]

    # 流式获取 AI 回复
    full_response = ""
    for partial in chat_with_ai(user_message, history_for_ai, username):
        full_response = partial
        display = chat_history + [{"role": "assistant", "content": partial}]
        yield display, session_id, "", gr.skip(), gr.skip()

    # 流式结束后持久化到 JSON
    persist_chat_message(username, session_id, user_message, full_response)

    # 刷新侧边栏数据
    df, ids = get_sessions_for_display(username)

    # 获取更新后的会话标题
    data = load_user_chat_history(username)
    title = "新对话"
    for s in data["sessions"]:
        if s["session_id"] == session_id:
            title = s.get("title", "新对话")
            break

    # 最终更新：完整的聊天记录 + 刷新侧边栏
    display = chat_history + [{"role": "assistant", "content": full_response}]
    yield display, session_id, "", df, ids


def handle_delete_click(username, session_id):
    """
    点击删除按钮 → 显示确认对话框
    返回 (confirm_text, delete_confirm_visible, delete_btn_visible, feedback_msg)
    """
    if not username or not session_id:
        return (
            "",                                              # delete_confirm_text
            gr.update(visible=False),                        # delete_confirm
            gr.update(visible=True),                         # delete_session_btn
            "❌ 请先选择一个会话"                             # chat_feedback
        )

    data = load_user_chat_history(username)
    title = "未知会话"
    for s in data["sessions"]:
        if s["session_id"] == session_id:
            title = s.get("title", "未知会话")
            break

    # 最后一个会话的特殊警告
    if len(data["sessions"]) <= 1:
        warning = "⚠️ 这是**最后一个会话**，删除后需新建会话才能继续聊天。"
    else:
        warning = "⚠️ 此操作**不可恢复**！"

    confirm_text = f"""### 🗑️ 确认删除会话？

{warning}

> 确定要删除会话「**{title}**」吗？
"""

    return (
        confirm_text,                        # delete_confirm_text
        gr.update(visible=True),             # delete_confirm 显示
        gr.update(visible=False),            # delete_session_btn 隐藏
        ""                                   # chat_feedback 清空
    )


def handle_confirm_delete(username, session_id):
    """
    确认删除 → 执行删除并更新全部UI
    返回 (session_id, chatbot, title_html, session_df, session_ids,
          confirm_visible, delete_btn_visible, feedback_msg)
    """
    if not username or not session_id:
        return (
            gr.skip(), gr.skip(), gr.skip(), gr.skip(), gr.skip(),
            gr.update(visible=False), gr.update(visible=True), "❌ 操作无效"
        )

    new_session_id, deleted_title = delete_chat_session(username, session_id)

    # 刷新侧边栏
    df, ids = get_sessions_for_display(username)

    if new_session_id:
        # 有剩余会话 → 切换到第一个可用会话
        msgs, title, sid = switch_session(username, new_session_id)
        return (
            sid,                                           # current_session_id
            msgs,                                          # chat_bot
            _make_session_title_html(title),               # session_title
            df,                                            # session_list
            ids,                                           # session_ids_state
            gr.update(visible=False),                      # delete_confirm 隐藏
            gr.update(visible=True),                       # delete_session_btn 显示
            f"✅ 已删除会话「{deleted_title}」"             # chat_feedback
        )
    else:
        # 无剩余会话 → 清空聊天区，提示用户新建
        return (
            "",                                            # current_session_id 清空
            [],                                            # chat_bot 清空
            _make_session_title_html("请新建会话"),         # session_title
            df,                                            # session_list（空DataFrame）
            ids,                                           # session_ids_state（空列表）
            gr.update(visible=False),                      # delete_confirm 隐藏
            gr.update(visible=True),                       # delete_session_btn 显示
            f"✅ 已删除会话「{deleted_title}」，请点击「➕ 新建聊天」开始新对话"  # chat_feedback
        )


def handle_cancel_delete():
    """
    取消删除 → 隐藏确认对话框，恢复删除按钮
    """
    return (
        gr.update(visible=False),    # delete_confirm 隐藏
        gr.update(visible=True),     # delete_session_btn 显示
        ""                           # chat_feedback 清空
    )


# -------------------------- 8. Gradio 界面搭建（多用户版）--------------------------

health_theme = gr.themes.Soft(
    primary_hue="emerald",
    secondary_hue="teal",
    neutral_hue="slate",
    font=[gr.themes.GoogleFont("Inter"), "ui-sans-serif", "system-ui", "Segoe UI", "Microsoft YaHei", "sans-serif"],
    radius_size=gr.themes.sizes.radius_lg,
    text_size=gr.themes.sizes.text_md
).set(
    button_primary_background_fill="*primary_500",
    button_primary_background_fill_hover="*primary_600",
    button_primary_text_color="white",
    button_secondary_background_fill="*background_fill_primary",
    button_secondary_background_fill_hover="*neutral_100",
    button_secondary_border_color="*neutral_200",
    button_secondary_text_color="*neutral_700",
    block_background_fill="*background_fill_primary",
    block_border_width="1px",
    block_border_color="*neutral_200",
    block_label_background_fill="transparent",
    block_label_border_width="0px",
    block_label_text_color="*neutral_600",
    block_label_text_weight="600",
    block_label_text_size="*text_sm",
    block_title_text_weight="600",
    block_title_text_color="*neutral_700",
    input_background_fill="*background_fill_primary",
    input_border_color="*neutral_200",
    input_border_color_focus="*primary_500",
    shadow_drop_lg="0 8px 24px rgba(15, 23, 42, 0.08)",
    block_shadow="0 1px 3px rgba(15, 23, 42, 0.05)"
)

custom_css = """
/* ==========================================================
   膳康管家 · 现代医疗健康风主题（Emerald 设计语言）
   设计规范：主色 #059669/#10b981 · 中性色 Slate · 圆角 10/16px
   阴影三级 · 全局微交互（hover 上浮 / focus 光环 / 过渡动画）
   ========================================================== */

/* ---------- 0. 全局基础 ---------- */
footer {display: none !important;}

.gradio-container {
    background:
        radial-gradient(1200px 500px at 85% -10%, rgba(16, 185, 129, 0.08), transparent 60%),
        radial-gradient(900px 420px at -10% 0%, rgba(20, 184, 166, 0.07), transparent 55%),
        #f6f8fb;
    font-feature-settings: "tnum";
    -webkit-font-smoothing: antialiased;
}

/* 细滚动条 */
::-webkit-scrollbar {width: 8px; height: 8px;}
::-webkit-scrollbar-track {background: transparent;}
::-webkit-scrollbar-thumb {background: #cbd5e1; border-radius: 8px;}
::-webkit-scrollbar-thumb:hover {background: #94a3b8;}

/* 组件卡片：白底 + 细边框 + 柔和阴影，悬停轻微增强 */
.gradio-container .block {
    border-radius: 12px !important;
    transition: box-shadow 0.25s ease, border-color 0.25s ease;
}

/* 进入动画 */
@keyframes fm-fade-up {
    from {opacity: 0; transform: translateY(10px);}
    to {opacity: 1; transform: translateY(0);}
}

/* ---------- 1. 按钮体系（微交互） ---------- */
.gradio-container button {
    transition: transform 0.18s ease, box-shadow 0.25s ease,
                background-color 0.2s ease, border-color 0.2s ease, filter 0.2s ease;
}
.gradio-container button:active {transform: translateY(1px) scale(0.99);}

/* 主按钮：渐变 + 悬停上浮 + 品牌色投影 */
.gradio-container button.primary,
.primary-btn button, button.primary-btn,
.new-chat-btn button, button.new-chat-btn,
.send-btn button, button.send-btn {
    background: linear-gradient(135deg, #059669, #10b981) !important;
    border: none !important;
    font-weight: 600 !important;
    box-shadow: 0 2px 6px rgba(5, 150, 105, 0.25) !important;
}
.gradio-container button.primary:hover,
.primary-btn button:hover, button.primary-btn:hover,
.new-chat-btn button:hover, button.new-chat-btn:hover,
.send-btn button:hover, button.send-btn:hover {
    transform: translateY(-2px);
    box-shadow: 0 10px 22px rgba(5, 150, 105, 0.35) !important;
    filter: brightness(1.04);
}

/* 次要按钮：白底描边，悬停浅绿 */
.gradio-container button.secondary:hover {
    border-color: #10b981 !important;
    color: #047857 !important;
    background: #ecfdf5 !important;
}

/* 危险按钮（variant=stop） */
.gradio-container button.stop:hover {filter: brightness(1.05);}

/* ---------- 2. 登录 / 注册页 ---------- */
#login_card_wrap {
    max-width: 460px;
    margin: 48px auto 24px auto;
    background: rgba(255, 255, 255, 0.92) !important;
    backdrop-filter: blur(10px);
    border: 1px solid rgba(226, 232, 240, 0.9) !important;
    border-radius: 20px !important;
    box-shadow: 0 20px 50px rgba(15, 23, 42, 0.10), 0 4px 14px rgba(5, 150, 105, 0.06) !important;
    padding: 36px 34px 28px 34px !important;
    animation: fm-fade-up 0.45s ease both;
}
/* 卡片内部组件块去除双层边框，形成整体感 */
#login_card_wrap .block {
    border: none !important;
    box-shadow: none !important;
    background: transparent !important;
}

.login-header {text-align: center; margin-bottom: 24px;}
.login-header h2 {border-left: none !important; padding-left: 0 !important;}
.login-header .fm-logo {
    width: 68px; height: 68px; margin: 0 auto 12px auto;
    display: flex; align-items: center; justify-content: center;
    font-size: 34px; border-radius: 20px;
    background: linear-gradient(135deg, #d1fae5, #a7f3d0);
    box-shadow: 0 8px 20px rgba(5, 150, 105, 0.22), inset 0 1px 0 rgba(255, 255, 255, 0.8);
}
.login-header h2 {
    color: #065f46; font-size: 1.7rem; font-weight: 800;
    margin: 0 0 6px 0; letter-spacing: 1px;
}
.login-header p {color: #64748b; font-size: 0.92rem; margin: 0;}

/* 登录输入框 */
.auth-input input, .auth-input textarea {
    border-radius: 10px !important;
    padding: 11px 14px !important;
    border: 1.5px solid #e2e8f0 !important;
    transition: border-color 0.2s ease, box-shadow 0.2s ease !important;
    font-size: 15px !important;
    background: #f8fafc !important;
}
.auth-input input:focus, .auth-input textarea:focus {
    border-color: #059669 !important;
    background: #ffffff !important;
    box-shadow: 0 0 0 4px rgba(5, 150, 105, 0.12) !important;
}
.auth-input span[data-testid="block-info"] {
    color: #475569 !important; font-weight: 600 !important; font-size: 13px !important;
}

/* 登录主按钮 */
.auth-btn button, button.auth-btn {
    border-radius: 10px !important;
    padding: 12px 24px !important;
    font-weight: 700 !important;
    font-size: 16px !important;
    letter-spacing: 4px;
}

/* 切换链接按钮（无边框透明链接样式） */
.switch-link button, button.switch-link {
    background: none !important;
    border: none !important;
    box-shadow: none !important;
    color: #059669 !important;
    font-size: 14px !important;
    font-weight: 500 !important;
    cursor: pointer !important;
    padding: 6px 10px !important;
    width: fit-content !important;
    margin: 2px auto 0 auto;
    white-space: nowrap !important;
}
.switch-link button:hover, button.switch-link:hover {
    background: none !important;
    color: #047857 !important;
    text-decoration: underline !important;
    transform: none;
}

/* 登录表单容器去灰底，融入卡片 */
#login_form, #register_form {
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
}
#login_card_wrap .form {
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
}

/* 提示消息条：空态隐藏，有内容时为浅绿消息条 */
.auth-msg {border: none !important; background: transparent !important; box-shadow: none !important;}
.auth-msg:has(textarea:placeholder-shown) {display: none;}
.auth-msg textarea, .auth-msg input {
    font-size: 14px !important;
    text-align: center !important;
    color: #065f46 !important;
    background: #ecfdf5 !important;
    border: 1px solid #a7f3d0 !important;
    border-radius: 10px !important;
    padding: 10px 12px !important;
    font-weight: 500;
    animation: fm-fade-up 0.3s ease both;
}

/* ---------- 3. 顶栏（用户信息 + 退出） ---------- */
#user_bar_row {
    align-items: center !important;
    gap: 12px;
    margin-bottom: 4px;
    flex-wrap: nowrap !important;
}
.fm-user-chip {
    display: flex; align-items: center; gap: 10px;
    padding: 8px 18px 8px 10px;
    background: linear-gradient(135deg, #ecfdf5, #d1fae5);
    border: 1px solid #a7f3d0;
    border-radius: 999px;
    box-shadow: 0 2px 8px rgba(5, 150, 105, 0.10);
    width: fit-content;
}
.fm-user-avatar {
    width: 32px; height: 32px; border-radius: 50%;
    background: linear-gradient(135deg, #059669, #10b981);
    color: #fff; font-weight: 700; font-size: 15px;
    display: flex; align-items: center; justify-content: center;
    box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.35);
}
.fm-user-name {color: #065f46; font-weight: 700; font-size: 14px; letter-spacing: 0.5px;}
.fm-user-dot {
    width: 8px; height: 8px; border-radius: 50%;
    background: #10b981; box-shadow: 0 0 0 3px rgba(16, 185, 129, 0.18);
}

.logout-btn button, button.logout-btn {
    border-radius: 999px !important;
    font-size: 13px !important;
    padding: 7px 18px !important;
    font-weight: 600 !important;
    white-space: nowrap !important;
    background: #ffffff !important;
    color: #dc2626 !important;
    border: 1px solid #fecaca !important;
    box-shadow: 0 1px 3px rgba(15, 23, 42, 0.06) !important;
}
.logout-btn button:hover, button.logout-btn:hover {
    background: #fef2f2 !important;
    border-color: #f87171 !important;
    box-shadow: 0 4px 12px rgba(220, 38, 38, 0.15) !important;
    transform: translateY(-1px);
}

/* ---------- 4. 主标题区 ---------- */
.fm-hero {text-align: center; max-width: 800px; margin: 0 auto; padding: 14px 0 22px 0; animation: fm-fade-up 0.4s ease both;}
.fm-hero h1, .fm-hero h3 {border-left: none !important; padding-left: 0 !important;}
.fm-hero h1 {
    color: #047857; font-size: 2.6rem; font-weight: 800;
    margin: 0 0 8px 0; letter-spacing: 2px;
    background: linear-gradient(135deg, #047857, #10b981);
    -webkit-background-clip: text; background-clip: text;
}
.fm-hero h3 {color: #475569; font-size: 1.15rem; font-weight: 400; margin: 0 0 14px 0;}
.fm-hero-badge {
    display: inline-block; background: #d1fae5; color: #065f46;
    padding: 5px 16px; border-radius: 999px;
    font-size: 0.85rem; font-weight: 500;
    border: 1px solid #a7f3d0;
}

/* ---------- 5. 标签页导航（胶囊式） ---------- */
.gradio-container div[role="tablist"] {
    gap: 6px;
    border-bottom: 1px solid #e2e8f0;
    padding-bottom: 6px;
    margin-bottom: 14px;
}
.gradio-container button[role="tab"] {
    border-radius: 999px !important;
    border: 1px solid transparent !important;
    background: transparent !important;
    color: #64748b !important;
    font-weight: 500 !important;
    font-size: 14px !important;
    padding: 8px 16px !important;
    margin: 0 !important;
    transition: all 0.22s ease !important;
}
.gradio-container button[role="tab"]:hover {
    background: #ecfdf5 !important;
    color: #047857 !important;
}
.gradio-container button[role="tab"][aria-selected="true"] {
    background: linear-gradient(135deg, #059669, #10b981) !important;
    color: #ffffff !important;
    font-weight: 600 !important;
    box-shadow: 0 4px 12px rgba(5, 150, 105, 0.28) !important;
    border-color: transparent !important;
}
/* 标签内容面板进入动画 */
.gradio-container div[role="tabpanel"] {animation: fm-fade-up 0.3s ease both;}

/* ---------- 6. Markdown 排版 ---------- */
.gradio-container .prose h2, .gradio-container .prose h3 {
    color: #0f172a !important;
    font-weight: 700 !important;
    padding-left: 12px;
    border-left: 4px solid #10b981;
    line-height: 1.35;
}
.gradio-container .prose hr {
    border: none; height: 1px;
    background: linear-gradient(90deg, transparent, #e2e8f0, transparent);
    margin: 20px 0;
}
.gradio-container .prose p {color: #334155;}

/* ---------- 7. 表单控件统一 ---------- */
.gradio-container input:not([type="checkbox"]):not([type="radio"]),
.gradio-container textarea {
    transition: border-color 0.2s ease, box-shadow 0.2s ease, background-color 0.2s ease !important;
}
.gradio-container input:focus, .gradio-container textarea:focus {
    box-shadow: 0 0 0 3px rgba(5, 150, 105, 0.12) !important;
}
/* 组件标签文字（去胶囊后统一为精致小标题） */
.gradio-container span[data-testid="block-info"] {letter-spacing: 0.3px;}

/* Radio 选项胶囊 */
.gradio-container label.svelte-1gzsjbx,
.gradio-container .wrap label {
    border-radius: 10px !important;
}
.gradio-container input[type="radio"] + * {
    transition: background-color 0.18s ease, border-color 0.18s ease, color 0.18s ease;
}

/* ---------- 8. 数据表格 ---------- */
.gradio-container [data-testid="dataframe"], .gradio-container .gradio-scheme > .wrap {
    border-radius: 12px !important;
    overflow: hidden;
}
.gradio-container table {border-collapse: separate; border-spacing: 0;}
.gradio-container thead tr th {
    background: #f0fdf4 !important;
    color: #065f46 !important;
    font-weight: 700 !important;
    font-size: 13px !important;
    border-bottom: 1px solid #d1fae5 !important;
}
.gradio-container tbody tr {transition: background-color 0.15s ease !important;}
.gradio-container tbody tr:nth-child(even) {background: #fafcfa;}
.gradio-container tbody tr:hover {background: #ecfdf5 !important;}

/* ---------- 9. 聊天区 ---------- */
#chat_sidebar {
    background: #ffffff;
    border-radius: 16px;
    padding: 14px;
    border: 1px solid #e2e8f0;
    box-shadow: 0 4px 16px rgba(15, 23, 42, 0.05);
    height: calc(100vh - 210px);
    overflow-y: auto;
}
#chat_sidebar h3, #chat_sidebar .prose h3 {
    color: #334155; font-size: 14px;
    margin-top: 14px; margin-bottom: 8px;
    border-left: none !important; padding-left: 0 !important;
}
.new-chat-btn button, button.new-chat-btn {
    border-radius: 10px !important;
    padding: 10px 20px !important;
    font-weight: 600 !important;
    font-size: 15px !important;
    width: 100% !important;
}

/* 会话列表 */
#session_table {font-size: 13px;}
#session_table table {border-collapse: collapse; width: 100%;}
#session_table thead tr th {background: #f8fafc !important; color: #64748b !important; border-bottom: 1px solid #e2e8f0 !important;}
#session_table tbody tr {
    cursor: pointer;
    transition: background-color 0.15s ease;
    border-radius: 8px;
}
#session_table tbody tr:hover {background-color: #d1fae5 !important;}
#session_table tbody tr.selected {
    background-color: #a7f3d0 !important;
    border-left: 3px solid #059669;
}

/* 主聊天区 */
#chat_main {
    background: #ffffff;
    border-radius: 16px;
    padding: 16px;
    border: 1px solid #e2e8f0;
    box-shadow: 0 4px 16px rgba(15, 23, 42, 0.05);
    height: calc(100vh - 210px);
    display: flex;
    flex-direction: column;
}

/* 会话标题条 */
.fm-session-title {
    padding: 10px 16px;
    background: linear-gradient(135deg, #ecfdf5, #d1fae5);
    border-radius: 12px;
    margin-bottom: 12px;
    border-left: 4px solid #059669;
    display: flex; align-items: center; gap: 8px;
}
.fm-session-title span {color: #065f46; font-weight: 600; font-size: 15px;}

/* 聊天气泡 */
#main_chatbot .message {
    font-size: 15px;
    line-height: 1.65;
    border-radius: 14px !important;
}
#main_chatbot {height: 100% !important; flex: 1;}

/* 聊天输入框 + 发送按钮 */
#chat_input textarea {
    border-radius: 12px !important;
    padding: 12px 16px !important;
    border: 1.5px solid #e2e8f0 !important;
    font-size: 15px !important;
    background: #f8fafc !important;
    transition: border-color 0.2s ease, box-shadow 0.2s ease, background-color 0.2s ease !important;
}
#chat_input textarea:focus {
    border-color: #059669 !important;
    background: #ffffff !important;
    box-shadow: 0 0 0 4px rgba(5, 150, 105, 0.10) !important;
}
.send-btn button, button.send-btn {
    border-radius: 12px !important;
    padding: 12px 20px !important;
    font-weight: 600 !important;
    height: 100%;
}

/* 聊天操作反馈条：空态隐藏 */
.chat-feedback {border: none !important; background: transparent !important; box-shadow: none !important;}
.chat-feedback:has(textarea:placeholder-shown) {display: none;}
.chat-feedback textarea, .chat-feedback input {
    font-size: 13px !important;
    text-align: center !important;
    color: #059669 !important;
    background: #ecfdf5 !important;
    border-radius: 8px !important;
    padding: 6px 10px !important;
    margin-top: 8px !important;
}

/* 删除当前会话按钮 */
.delete-session-btn button, button.delete-session-btn {
    border-radius: 8px !important;
    font-size: 13px !important;
    width: 100% !important;
    margin-top: 8px !important;
    background: #fef2f2 !important;
    color: #dc2626 !important;
    border: 1px solid #fecaca !important;
    transition: all 0.2s ease !important;
}
.delete-session-btn button:hover, button.delete-session-btn:hover {
    background: #fee2e2 !important;
    border-color: #f87171 !important;
}

/* 确认删除对话框 */
#delete_confirm {
    background: #fef2f2;
    border: 1.5px solid #fecaca;
    border-radius: 12px;
    padding: 14px;
    margin-top: 10px;
    animation: fm-fade-up 0.25s ease both;
}
#delete_confirm h3 {color: #dc2626; font-size: 15px; margin-top: 0; margin-bottom: 8px;}
#delete_confirm p {font-size: 13px; color: #7f1d1d; margin-bottom: 6px;}
#delete_confirm blockquote {
    border-left: 3px solid #f87171;
    padding-left: 10px;
    margin: 8px 0;
    color: #991b1b;
    font-size: 13px;
}

/* ---------- 10. 分组卡片（记录操作区 / 趋势区） ---------- */
.fm-panel {
    background: #ffffff;
    border: 1px solid #e2e8f0;
    border-radius: 14px;
    padding: 14px 14px 6px 14px;
    box-shadow: 0 2px 10px rgba(15, 23, 42, 0.04);
    margin: 6px 0;
}

/* 查房小结输出框高度收敛 */
.fm-panel .block:has(> label > span[data-testid="block-info"]) textarea {min-height: 120px;}
"""
with gr.Blocks(title="膳康管家 - 智能医疗系统") as demo:
    # ===== 会话状态 =====
    current_user = gr.State("")
    current_session_id = gr.State("")
    session_ids_state = gr.State([])

    # ===== 登录/注册区块 =====
    with gr.Column(visible=True, elem_id="login_card_wrap") as login_block:
        # 精美头部
        gr.HTML("""
        <div class="login-header">
            <div class="fm-logo">🥗</div>
            <h2>膳康管家</h2>
            <p>个性化膳食图像识别与营养智能管理系统</p>
        </div>
        """)

        # 提示消息
        auth_message = gr.Textbox(
            label="", show_label=False, interactive=False,
            elem_classes="auth-msg", container=True
        )

        # ---- 登录表单 ----
        with gr.Column(visible=True, elem_id="login_form") as login_form:
            login_username = gr.Textbox(
                label="👤 用户名", placeholder="请输入用户名",
                elem_classes="auth-input"
            )
            login_password = gr.Textbox(
                label="🔒 密码", placeholder="请输入密码", type="password",
                elem_classes="auth-input"
            )
            with gr.Row():
                login_btn = gr.Button("🔑 登 录", variant="primary", elem_classes="auth-btn", scale=1)
            with gr.Row():
                to_register_btn = gr.Button("没有账号？立即注册 →", elem_classes="switch-link", scale=0)

        # ---- 注册表单 ----
        with gr.Column(visible=False, elem_id="register_form") as register_form:
            reg_username = gr.Textbox(
                label="👤 用户名", placeholder="请输入用户名（字母和数字）",
                elem_classes="auth-input"
            )
            reg_password = gr.Textbox(
                label="🔒 密码", placeholder="请输入密码（至少6位）", type="password",
                elem_classes="auth-input"
            )
            reg_confirm = gr.Textbox(
                label="🔒 确认密码", placeholder="请再次输入密码", type="password",
                elem_classes="auth-input"
            )
            with gr.Row():
                reg_btn = gr.Button("📝 注 册", variant="primary", elem_classes="auth-btn", scale=1)
            with gr.Row():
                to_login_btn = gr.Button("← 已有账号？返回登录", elem_classes="switch-link", scale=0)

    # ===== 主应用区块 =====
    with gr.Column(visible=False) as main_block:
        # 用户信息栏 + 退出按钮
        with gr.Row(elem_id="user_bar_row"):
            user_bar = gr.HTML("")
            logout_btn = gr.Button("🚪 退出登录", variant="stop", elem_classes="logout-btn", scale=0, min_width=100)

        # 应用标题
        gr.HTML("""
        <div class="fm-hero">
            <h1>🥗 膳康管家</h1>
            <h3>个性化膳食图像识别与营养智能管理系统</h3>
            <span class="fm-hero-badge">4C中国大学生计算机设计大赛参赛作品 | AI+智能医疗健康方向</span>
        </div>
        """)

        with gr.Tab("📊 今日概览", id="dashboard"):
            gr.Markdown("### 📊 今日营养概览")
            with gr.Row():
                dash_refresh_btn = gr.Button("🔄 刷新概览", variant="primary", elem_classes="primary-btn")
            with gr.Row():
                dash_text = gr.Textbox(label="今日摄入摘要", lines=10, scale=1)
                dash_progress = gr.Image(label="营养指标完成度", scale=1)
            gr.Markdown("---")
            gr.Markdown("💡 **快捷入口**：切换到上方标签页进行 [📸 识别食物] / [📊 查看记录] / [💬 AI咨询]")

        with gr.Tab("📸 食品识别与营养解析"):
            gr.Markdown("### 📸 图像识别模式")
            with gr.Row():
                with gr.Column(scale=1):
                    image_input = gr.Image(type="pil", label="📸 上传美食/水果图片（支持多目标同框识别）")
                    weight_input = gr.Number(value=100, label="总食用重量(g)，默认100g时使用系统估算值", minimum=1)
                    date_input = gr.Textbox(label="用餐日期", value=datetime.date.today().isoformat(), placeholder="建议格式：YYYY-MM-DD")
                    meal_type_input = gr.Radio(
                        ["早餐", "午餐", "晚餐", "加餐"],
                        label="餐次类型",
                        value="午餐"
                    )
                    predict_btn = gr.Button("🔍 一键识别与解析", variant="primary", elem_classes="primary-btn")
                with gr.Column(scale=1):
                    annotated_output = gr.Image(label="识别结果（带检测框）")
                    nutrition_output = gr.Textbox(label="营养成分详情", lines=10)
                    record_output = gr.Textbox(label="记录状态", lines=2)

            # 手动录入模式
            gr.Markdown("---")
            gr.Markdown("### ✍️ 手动录入模式（无需照片）")
            with gr.Row():
                with gr.Column(scale=1):
                    manual_food_input = gr.Dropdown(
                        choices=FOOD_CLASSES,
                        multiselect=True,
                        label="选择食物（可多选）",
                        info="从下拉列表中选择一种或多种食物，支持搜索"
                    )
                    manual_weight_input = gr.Textbox(
                        label="对应重量(g)（逗号分隔，留空则使用默认份量）",
                        placeholder="如：200, 150, 200（按选择顺序填写，个数需与食物一致）",
                        value=""
                    )
                    manual_date_input = gr.Textbox(label="用餐日期", value=datetime.date.today().isoformat(), placeholder="YYYY-MM-DD")
                    manual_meal_input = gr.Radio(
                        ["早餐", "午餐", "晚餐", "加餐"],
                        label="餐次类型",
                        value="午餐"
                    )
                    manual_add_btn = gr.Button("✍️ 手动添加记录", variant="secondary", elem_classes="primary-btn")
                with gr.Column(scale=1):
                    manual_nutrition_output = gr.Textbox(label="营养成分", lines=8)
                    manual_record_output = gr.Textbox(label="记录状态", lines=2)

        with gr.Tab("👤 个人档案与推荐设置"):
            with gr.Row():
                with gr.Column(scale=1):
                    name_input = gr.Textbox(label="姓名", value="用户")
                    height_input = gr.Number(label="身高(cm)", value=170, minimum=100)
                    profile_weight_input = gr.Number(label="体重(kg)", value=65, minimum=30)
                    age_input = gr.Number(label="年龄", value=20, minimum=10)
                with gr.Column(scale=1):
                    gender_input = gr.Radio(["男", "女"], label="性别", value="男")
                    activity_input = gr.Radio(
                        ["久坐不动", "轻度活动", "中度活动", "重度活动", "极重度活动"],
                        label="日常活动量", value="轻度活动"
                    )
                    goal_input = gr.Radio(
                        ["减脂", "维持体重", "增肌", "糖尿病管理", "高血压管理"],
                        label="健康目标", value="维持体重"
                    )
                    disease_input = gr.Radio(
                        ["健康", "糖尿病", "高血压", "高血脂"],
                        label="身体状况", value="健康"
                    )
            profile_btn = gr.Button("✅ 更新个人档案", variant="primary", elem_classes="primary-btn")
            profile_output = gr.Textbox(label="个人档案与推荐摄入量", lines=12)

        with gr.Tab("📊 膳食记录与健康报告"):
            with gr.Row():
                query_date = gr.Textbox(label="查询日期", value=datetime.date.today().isoformat(), placeholder="YYYY-MM-DD")
                query_btn = gr.Button("🔍 查询当日记录", variant="primary", elem_classes="primary-btn")
                export_btn = gr.Button("📥 导出Excel报告", variant="secondary")
            with gr.Row():
                report_output = gr.Textbox(label="膳食营养统计报告", lines=18, scale=2)
                progress_ring_output = gr.Image(label="营养指标完成度（环形图）", scale=1)
            records_output = gr.Dataframe(label="当日膳食明细记录", wrap=True)

            # 删除记录功能
            with gr.Row(elem_classes="fm-panel"):
                delete_index_input = gr.Number(value=0, label="要删除的记录序号（从0开始，参考上方表格行号）", precision=0, minimum=0)
                delete_btn = gr.Button("🗑️ 删除选中记录", variant="stop")
                delete_status = gr.Textbox(label="删除状态", lines=1)
            export_status = gr.Textbox(label="导出状态", lines=1)

            gr.Markdown("### 📈 本周趋势分析")
            with gr.Row(elem_classes="fm-panel"):
                weekly_btn = gr.Button("查看本周多维度趋势", variant="primary")
                weekly_msg = gr.Textbox(label="统计信息", lines=1)
                weekly_chart = gr.Image(label="营养摄入趋势图（热量/蛋白质/碳水/脂肪）")

        # AI 专属健康咨询标签页（自定义聊天界面，参考 Ollama 风格）
        with gr.Tab("💬 AI 专属健康咨询"):
            gr.Markdown("### 🤖 您的私人营养师（三甲医院临床营养科主任医师人设）已上线")
            with gr.Row(elem_classes="fm-panel"):
                daily_summary_btn = gr.Button("🩺 查看主任医师今日查房小结", variant="secondary")
                daily_summary_output = gr.Textbox(label="📋 今日查房小结", lines=12)
            gr.Markdown("---")

            # 左右两栏布局：侧边栏 + 主聊天区
            with gr.Row():
                # ---- 左侧边栏：会话管理 ----
                with gr.Column(scale=1, min_width=220, elem_id="chat_sidebar"):
                    new_chat_btn = gr.Button(
                        "➕ 新建聊天", variant="primary",
                        elem_classes="new-chat-btn"
                    )
                    gr.Markdown("### 💬 历史会话")
                    session_list = gr.Dataframe(
                        headers=["💬 会话", "🕒 更新时间"],
                        value=pd.DataFrame({"💬 会话": [], "🕒 更新时间": []}),
                        elem_id="session_table",
                        wrap=True,
                    )

                    # 操作反馈消息
                    chat_feedback = gr.Textbox(
                        label="", show_label=False, interactive=False,
                        elem_classes="chat-feedback", container=True,
                        placeholder=""
                    )

                    # 删除当前会话按钮
                    delete_session_btn = gr.Button(
                        "🗑️ 删除当前会话", variant="stop",
                        elem_classes="delete-session-btn",
                        size="sm"
                    )

                    # 确认删除对话框（默认隐藏）
                    with gr.Column(visible=False, elem_id="delete_confirm") as delete_confirm:
                        delete_confirm_text = gr.Markdown("")
                        with gr.Row():
                            confirm_delete_btn = gr.Button(
                                "✅ 确认删除", variant="stop",
                                size="sm", scale=1
                            )
                            cancel_delete_btn = gr.Button(
                                "❌ 取消", size="sm", scale=1
                            )

                # ---- 右侧主聊天区 ----
                with gr.Column(scale=4, elem_id="chat_main"):
                    session_title = gr.HTML(
                        value="""<div style="padding: 10px 16px; background: linear-gradient(135deg, #ecfdf5, #d1fae5); border-radius: 10px; margin-bottom: 10px; border-left: 4px solid #059669;">
                        <span style="color: #065f46; font-weight: 600; font-size: 15px;">💬 新对话</span>
                        </div>"""
                    )
                    chat_bot = gr.Chatbot(
                        value=[],
                        height=500,
                        avatar_images=(
                            os.path.join(PICTURE_DIR, "boy.jpg"),
                            os.path.join(PICTURE_DIR, "doctor.jpg")
                        ),
                        elem_id="main_chatbot",
                    )
                    with gr.Row():
                        msg_input = gr.Textbox(
                            placeholder="请输入您的健康问题...（Enter 发送）",
                            container=False,
                            scale=7,
                            elem_id="chat_input",
                        )
                        send_btn = gr.Button("📨 发送", variant="primary", scale=1, elem_classes="send-btn")

            gr.Markdown("您可以咨询：'我今天蛋白质吃够了吗？'、'晚上去健身，现在适合吃根香蕉吗？'、'高血压患者能吃识别出来的这个菜吗？'")

    # ============ 事件绑定 ============

    # ---- 登录区块事件 ----

    # 登录按钮 → 初始化聊天状态 → 更新头像 → 加载档案
    login_btn.click(
        fn=handle_login,
        inputs=[login_username, login_password],
        outputs=[current_user, login_block, main_block, login_form, register_form, auth_message]
    ).then(
        fn=init_chat_state,
        inputs=[current_user],
        outputs=[current_session_id, session_list, session_ids_state, session_title, chat_bot]
    ).then(
        fn=update_chat_avatar,
        inputs=[current_user],
        outputs=[chat_bot]
    ).then(
        fn=load_profile_to_form,
        inputs=[current_user],
        outputs=[name_input, height_input, profile_weight_input, age_input,
                 gender_input, activity_input, goal_input, disease_input]
    )

    # 注册按钮 → 初始化聊天状态 → 更新头像 → 加载档案
    reg_btn.click(
        fn=handle_register,
        inputs=[reg_username, reg_password, reg_confirm],
        outputs=[current_user, login_block, main_block, login_form, register_form, auth_message]
    ).then(
        fn=init_chat_state,
        inputs=[current_user],
        outputs=[current_session_id, session_list, session_ids_state, session_title, chat_bot]
    ).then(
        fn=update_chat_avatar,
        inputs=[current_user],
        outputs=[chat_bot]
    ).then(
        fn=load_profile_to_form,
        inputs=[current_user],
        outputs=[name_input, height_input, profile_weight_input, age_input,
                 gender_input, activity_input, goal_input, disease_input]
    )

    # 切换到注册表单
    to_register_btn.click(
        fn=switch_to_register,
        inputs=None,
        outputs=[login_form, register_form, auth_message]
    )

    # 切换到登录表单
    to_login_btn.click(
        fn=switch_to_login,
        inputs=None,
        outputs=[login_form, register_form, auth_message]
    )

    # ---- 退出登录 → 清空聊天状态 → 重置头像 ----
    logout_btn.click(
        fn=handle_logout,
        inputs=None,
        outputs=[current_user, login_block, main_block, login_form, register_form, auth_message]
    ).then(
        fn=lambda: ("", pd.DataFrame({"💬 会话": [], "🕒 更新时间": []}), [], "", []),
        inputs=None,
        outputs=[current_session_id, session_list, session_ids_state, session_title, chat_bot]
    ).then(
        fn=update_chat_avatar,
        inputs=[current_user],  # 此时 current_user 已由 handle_logout 设为空字符串
        outputs=[chat_bot]
    ).then(
        fn=load_profile_to_form,
        inputs=[current_user],  # 空字符串 → 重置为默认值
        outputs=[name_input, height_input, profile_weight_input, age_input,
                 gender_input, activity_input, goal_input, disease_input]
    )

    # 登录成功时更新用户信息栏（通过 current_user 变化触发）
    current_user.change(
        fn=get_user_bar_html,
        inputs=[current_user],
        outputs=[user_bar]
    )

    # ---- 主应用功能事件（全部加上 current_user 作为第一个输入）----

    # 仪表盘刷新
    dash_refresh_btn.click(
        fn=create_dashboard,
        inputs=[current_user],
        outputs=[dash_text, dash_progress]
    )

    # 图像识别
    predict_btn.click(
        fn=main_predict,
        inputs=[current_user, image_input, weight_input, date_input, meal_type_input],
        outputs=[nutrition_output, record_output, annotated_output]
    )

    # 手动录入
    manual_add_btn.click(
        fn=manual_add_food,
        inputs=[current_user, manual_food_input, manual_weight_input, manual_date_input, manual_meal_input],
        outputs=[manual_nutrition_output, manual_record_output]
    )

    # 个人档案更新 → 同步刷新头像 → 刷新输入框
    profile_btn.click(
        fn=update_profile_handler,
        inputs=[current_user, name_input, height_input, profile_weight_input, age_input,
                gender_input, activity_input, goal_input, disease_input],
        outputs=profile_output
    ).then(
        fn=update_chat_avatar,
        inputs=[current_user],
        outputs=[chat_bot]
    ).then(
        fn=load_profile_to_form,
        inputs=[current_user],
        outputs=[name_input, height_input, profile_weight_input, age_input,
                 gender_input, activity_input, goal_input, disease_input]
    )

    # 查询记录
    query_btn.click(
        fn=query_records,
        inputs=[current_user, query_date],
        outputs=[report_output, records_output, progress_ring_output]
    )

    # 删除记录
    delete_btn.click(
        fn=delete_record_handler,
        inputs=[current_user, query_date, delete_index_input],
        outputs=[delete_status, report_output, records_output, progress_ring_output]
    )

    # 导出Excel
    export_btn.click(
        fn=export_report,
        inputs=[current_user, query_date],
        outputs=export_status
    )

    # 周趋势
    weekly_btn.click(
        fn=show_weekly_trend,
        inputs=[current_user],
        outputs=[weekly_msg, weekly_chart]
    )

    # AI每日小结
    daily_summary_btn.click(
        fn=get_daily_summary_for_ai,
        inputs=[current_user],
        outputs=daily_summary_output
    )

    # ---- 聊天界面事件 ----

    # 新建聊天
    new_chat_btn.click(
        fn=handle_new_chat,
        inputs=[current_user],
        outputs=[current_session_id, chat_bot, session_title, session_list, session_ids_state]
    )

    # 点击侧边栏会话条目切换会话
    session_list.select(
        fn=handle_session_select,
        inputs=[current_user, session_ids_state],
        outputs=[current_session_id, chat_bot, session_title, msg_input, session_list, session_ids_state]
    )

    # 发送消息（流式生成器）
    send_btn.click(
        fn=handle_send_message,
        inputs=[current_user, current_session_id, msg_input, chat_bot],
        outputs=[chat_bot, current_session_id, msg_input, session_list, session_ids_state]
    )

    # 回车发送（通过 msg_input 的 submit 事件）
    msg_input.submit(
        fn=handle_send_message,
        inputs=[current_user, current_session_id, msg_input, chat_bot],
        outputs=[chat_bot, current_session_id, msg_input, session_list, session_ids_state]
    )

    # ---- 删除会话事件 ----

    # 点击删除按钮 → 显示确认对话框
    delete_session_btn.click(
        fn=handle_delete_click,
        inputs=[current_user, current_session_id],
        outputs=[delete_confirm_text, delete_confirm, delete_session_btn, chat_feedback]
    )

    # 确认删除 → 执行删除
    confirm_delete_btn.click(
        fn=handle_confirm_delete,
        inputs=[current_user, current_session_id],
        outputs=[
            current_session_id, chat_bot, session_title,
            session_list, session_ids_state,
            delete_confirm, delete_session_btn, chat_feedback
        ]
    )

    # 取消删除 → 隐藏确认框
    cancel_delete_btn.click(
        fn=handle_cancel_delete,
        inputs=None,
        outputs=[delete_confirm, delete_session_btn, chat_feedback]
    )


if __name__ == "__main__":
    logger.info("膳康管家多用户系统启动中...")
    demo.launch(inbrowser=True, share=False, theme=health_theme, css=custom_css)
